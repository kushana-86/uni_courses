import re
import time
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).parent
TARGET = 26954
RANK_URL = "https://www.dxsbb.com/news/148772.html"


def load_rank_map():
    last = None
    for attempt in range(5):
        try:
            table = pd.read_html(RANK_URL)[0]
            break
        except Exception as exc:
            last = exc
            time.sleep(attempt + 1)
    else:
        raise last
    table.columns = ["score", "count", "rank"]
    out = {}
    for _, row in table.iterrows():
        m = re.search(r"\d+", str(row["score"]))
        if m:
            try:
                out[int(m.group())] = int(float(row["rank"]))
            except (TypeError, ValueError):
                pass
    return out


def value(major):
    weights = {
        "计算机": 13, "软件": 12, "电气": 12, "电子信息": 11, "自动化": 11,
        "人工智能": 11, "数据科学": 10, "通信": 10, "信息工程": 10,
        "机械": 8, "数学": 8, "统计": 8, "新能源": 8, "能源": 7,
        "临床医学": 9, "医学技术": 7, "药学": 6, "法学": 7,
        "经济": 6, "金融": 6, "地理": 6, "水利": 6, "材料": 5,
        "环境": 4, "化学": 5, "土木": 2, "矿业": 2, "护理": 1,
    }
    return max((v for k, v in weights.items() if k in major), default=3)


rank_map = load_rank_map()
rows = []
for raw in (ROOT / "data" / "2025_B_layout.txt").read_text(encoding="utf-8").splitlines():
    line = " ".join(raw.split())
    if "高校专项" not in line:
        continue
    m = re.match(r"^(\d{4})\s*(.+?\(高校专项\))\s*([0-9A-Z]{3})\s*(.+?)\s*(\d{3})(?:\s|$)", line)
    if not m:
        continue
    code, school, major_code, major, score = m.groups()
    score = int(score)
    rank = rank_map.get(score)
    if not rank or not (8000 <= rank <= 75000):
        continue
    if "思想政治" in major:
        subject = "物化地通常不适配；2026计划再核"
    else:
        subject = "物化地初筛可考虑；2026计划再核"
    rows.append({
        "code": code, "school": school, "major_code": major_code, "major": major,
        "score": score, "rank": rank, "subject": subject,
    })

# 60个B段专业平行志愿：拉开梯度，并控制同一院校最多3项。
segments = [(8000, 20000, 15), (20000, 24000, 10), (24000, 30000, 18),
            (30000, 45000, 11), (45000, 75001, 6)]
selected, school_count = [], {}
for low, high, count in segments:
    pool = [r for r in rows if low <= r["rank"] < high]
    pool.sort(key=lambda r: (-value(r["major"]), abs(r["rank"] - TARGET)))
    added = 0
    for row in pool:
        school_key = re.sub(r"\(.*?\)", "", row["school"])
        if school_count.get(school_key, 0) >= 6:
            continue
        school_count[school_key] = school_count.get(school_key, 0) + 1
        selected.append(row)
        added += 1
        if added == count:
            break
selected.sort(key=lambda r: r["rank"])

wb = Workbook()
ws = wb.active
ws.title = "高校专项60项方案"
headers = [
    "建议序号", "院校代码", "院校", "专业代码", "专业", "物化地适配初筛",
    "2022最低位次", "2023最低位次", "2024最低位次", "2025最低分",
    "2025最低位次", "与本人26954名差值", "梯度", "专业优先度", "填报前提",
]
ws.append(headers)
for i, row in enumerate(selected, 1):
    diff = row["rank"] - TARGET
    band = "冲" if row["rank"] < 24000 else ("稳" if row["rank"] <= 32000 else "保")
    ws.append([
        i, row["code"], re.sub(r"\(高校专项\)", "", row["school"]),
        row["major_code"], re.sub(r"\(高校专项.*?\)", "", row["major"]),
        row["subject"],
        "无同口径数据", "无同口径数据", "无同口径数据",
        row["score"], row["rank"], diff, band, value(row["major"]),
        "仅限已通过高校专项资格及相应高校审核；以2026正式计划为准",
    ])

tab = Table(displayName="UniversitySpecialPlan", ref=f"A1:O{ws.max_row}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tab)
ws.freeze_panes = "G2"
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="5B2C6F")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    row[12].fill = PatternFill("solid", fgColor={
        "冲": "F4CCCC", "稳": "FFF2CC", "保": "D9EAD3"
    }[row[12].value])
widths = [10, 12, 28, 12, 42, 28, 16, 16, 16, 13, 15, 18, 9, 13, 55]
for i, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

guide = wb.create_sheet("填报策略")
strategy = [
    ("项目", "建议"),
    ("考生基准", "重庆物理类，物化地，566分，26954名。"),
    ("使用前提", "必须已经通过2026高校专项报考资格，并通过具体高校的高校专项审核。未通过的高校不能填。"),
    ("数量", "按本科提前批B段专业平行志愿思路整理60项。正式可填数量和专业以2026志愿系统显示为准。"),
    ("排序", "先删除未通过高校审核、选科不符或不能接受的专业，再按真实专业意愿排序；不要仅按分数从高到低机械排列。"),
    ("梯度", "冲：2025位次优于24000；稳：24000—32000；保：32000以后。高校专项样本小、年度波动大，梯度只作参考。"),
    ("2022—2024", "2025年重庆将高校专项调整到本科提前批B段并实行平行志愿。此前年度不是同一专业平行投档口径，故不填伪可比位次。"),
    ("2025", "采用重庆2025本科提前批B段物理类专业投档最低分，并用当年物理类一分一段表换算位次。"),
    ("普通本科批", "高校专项未录取时仍进入后续普通本科批；请继续保留此前制作的普通本科批100项方案。"),
]
for row in strategy:
    guide.append(row)
guide.column_dimensions["A"].width = 20
guide.column_dimensions["B"].width = 115
for cell in guide[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="5B2C6F")
for row in guide.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

src = wb.create_sheet("来源")
sources = [
    ("资料", "链接/说明"),
    ("2025高校专项专业投档数据", "https://cdn.gaokzx.com/zixunzhan/1753263227044%E9%87%8D%E5%BA%86%E5%B8%82%E6%99%AE%E9%80%9A%E9%AB%98%E6%A0%A1%E6%8B%9B%E7%94%9F%E4%BF%A1%E6%81%AF%E8%A1%A8%E6%9C%AC%E7%A7%91%E6%8F%90%E5%89%8D%E6%89%B9B%E6%AE%B5-%E7%89%A9%E7%90%86%E7%B1%BB.pdf"),
    ("2025物理类一分一段表", RANK_URL),
    ("2026重庆招生实施办法", "https://www.gkzxw.com/Article/202605/73458.html"),
    ("口径变化说明", "2025高校专项调整至本科提前批B段并实行平行志愿；早年数据不直接横比。"),
]
for row in sources:
    src.append(row)
for r in range(2, src.max_row + 1):
    val = src.cell(r, 2).value
    if isinstance(val, str) and val.startswith("http"):
        src.cell(r, 2).hyperlink = val
        src.cell(r, 2).style = "Hyperlink"
src.column_dimensions["A"].width = 30
src.column_dimensions["B"].width = 115
for cell in src[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="5B2C6F")

out = ROOT / "2026重庆高考_仅高校专项_60项填报方案.xlsx"
wb.save(out)
print(out)
print("parsed:", len(rows), "selected:", len(selected))
