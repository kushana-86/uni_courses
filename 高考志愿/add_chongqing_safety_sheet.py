import re
import statistics
import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).parent
BOOK = ROOT / "2026重庆高考_提前批AB段加本科批_整体填报方案.xlsx"
TARGET = 26954
RANK_URLS = {
    2022: "https://www.dxsbb.com/news/117895.html",
    2023: "https://www.dxsbb.com/news/136970.html",
    2024: "https://www.dxsbb.com/news/146467.html",
    2025: "https://www.dxsbb.com/news/148772.html",
}
PUBLIC = {
    "重庆邮电大学", "重庆交通大学", "重庆医科大学", "重庆师范大学", "重庆文理学院",
    "重庆三峡学院", "长江师范学院", "重庆科技大学", "重庆理工大学", "重庆工商大学",
    "四川外国语大学", "重庆第二师范学院", "重庆中医药学院", "四川美术学院",
}
PRIVATE = {
    "重庆人文科技学院", "重庆外语外事学院", "重庆对外经贸学院", "重庆财经学院",
    "重庆工商大学派斯学院", "重庆移通学院", "重庆城市科技学院", "重庆工程学院",
    "重庆机电职业技术大学",
}
FEATURES = {
    "重庆邮电大学": "信息通信、计算机、自动化",
    "重庆交通大学": "交通运输、土木、水利、车辆",
    "重庆医科大学": "临床医学、公共卫生、医学技术",
    "重庆师范大学": "教育学、数学、汉语言文学",
    "重庆文理学院": "材料、机械、师范教育、园林",
    "重庆三峡学院": "水利电力、电子信息、师范教育",
    "长江师范学院": "师范教育、化学、材料、环境",
    "重庆科技大学": "石油工程、冶金材料、机械、安全工程",
    "重庆理工大学": "车辆工程、机械、材料、会计",
    "重庆工商大学": "工商管理、应用经济、食品科学",
    "四川外国语大学": "外国语言文学、国际传播",
    "重庆第二师范学院": "师范教育、食品质量与安全、旅游管理",
    "重庆中医药学院": "中医学、中药学",
}


def rank_map(year):
    last = None
    for attempt in range(5):
        try:
            table = pd.read_html(RANK_URLS[year])[0]
            break
        except Exception as exc:
            last = exc
            time.sleep(attempt + 1)
    else:
        raise last
    table.columns = ["score", "count", "rank"]
    out = {}
    for _, row in table.iterrows():
        m = re.search(r"\d+", str(row["score"]))
        if m:
            try:
                out[int(m.group())] = int(float(row["rank"]))
            except (TypeError, ValueError):
                pass
    return out


def clean_school(name):
    return re.sub(r"\((地方专项|国家专项|高校专项|中外合作|民族班|预科班).*?\)", "", name).strip()


def clean_major(name):
    name = re.sub(r"\(.*?\)", "", name)
    return re.sub(r"(类|试验班.*)$", "", name).strip()


ranks = {year: rank_map(year) for year in range(2022, 2026)}
year_rows = {}
for year in range(2022, 2026):
    data = []
    for raw in (ROOT / "data" / f"{year}.txt").read_text(encoding="utf-8").splitlines():
        line = " ".join(raw.split())
        m = re.match(r"^(\d{4})\s+(.+?)\s+([0-9A-Z]{3})\s+(.+?)\s+(\d{3})(?:\s|$)", line)
        if not m:
            continue
        code, school, major_code, major, score = m.groups()
        base_school = clean_school(school)
        if base_school not in PUBLIC and base_school not in PRIVATE:
            continue
        if any(x in school for x in ("地方专项", "国家专项", "高校专项", "民族班", "预科班", "中外合作")):
            continue
        score = int(score)
        rank = ranks[year].get(score)
        if rank:
            data.append({
                "code": code, "school": base_school, "major_code": major_code,
                "major": major, "major_key": clean_major(major), "score": score, "rank": rank,
            })
    year_rows[year] = data

index = {}
for year, rows in year_rows.items():
    for row in rows:
        index.setdefault((year, row["school"], row["major_key"]), []).append(row)

results, seen = [], set()
for current in year_rows[2025]:
    key = (current["school"], current["major_key"])
    if key in seen or "思想政治教育" in current["major"]:
        continue
    seen.add(key)
    history = {}
    for year in range(2022, 2026):
        matches = index.get((year, *key), [])
        if matches:
            match = sorted(matches, key=lambda r: len(r["major"]))[0]
            history[year] = (match["score"], match["rank"])
        else:
            history[year] = (None, None)
    valid = [rank for _, rank in history.values() if rank]
    if len(valid) < 2:
        continue
    median = round(statistics.median(valid))
    if current["school"] in PUBLIC and median < 30000 and current["rank"] < 33000:
        continue
    if current["school"] in PRIVATE:
        category = "民办最终兜底"
    elif median < 38000:
        category = "准保底（仍有波动）"
    elif median < 55000:
        category = "公办稳妥保底"
    else:
        category = "公办深度保底"
    results.append({
        **current, "history": history, "median": median, "best": min(valid),
        "worst": max(valid), "volatility": max(valid) - min(valid),
        "years": len(valid), "category": category,
    })

order = {"准保底（仍有波动）": 1, "公办稳妥保底": 2, "公办深度保底": 3, "民办最终兜底": 4}
results.sort(key=lambda r: (order[r["category"]], r["median"], r["school"], r["major"]))

wb = load_workbook(BOOK)
if "重庆院校保底" in wb.sheetnames:
    del wb["重庆院校保底"]
ws = wb.create_sheet("重庆院校保底", 3)
headers = [
    "序号", "保底层级", "办学性质", "院校", "专业", "院校特色方向",
    "2022最低分", "2022最低位次", "2023最低分", "2023最低位次",
    "2024最低分", "2024最低位次", "2025最低分", "2025最低位次",
    "近年位次中位数", "最好位次", "最差位次", "波动幅度", "有效年份",
    "相对本人位次差", "填报提示",
]
ws.append(headers)
for i, row in enumerate(results, 1):
    h = row["history"]
    nature = "公办" if row["school"] in PUBLIC else "民办/职业本科"
    tip = "可作为重庆本地保底，但仍须核验2026计划、选科和体检"
    if nature != "公办":
        tip = "最终兜底；重点核验学费、校区、转专业和培养模式"
    ws.append([
        i, row["category"], nature, row["school"], row["major"],
        FEATURES.get(row["school"], "应用型专业方向，具体优势以学校官网为准"),
        h[2022][0], h[2022][1], h[2023][0], h[2023][1],
        h[2024][0], h[2024][1], h[2025][0], h[2025][1],
        row["median"], row["best"], row["worst"], row["volatility"], row["years"],
        row["median"] - TARGET, tip,
    ])

for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="006100")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
colors = {
    "准保底（仍有波动）": "FFF2CC", "公办稳妥保底": "D9EAD3",
    "公办深度保底": "B6D7A8", "民办最终兜底": "D9D2E9",
}
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    row[1].fill = PatternFill("solid", fgColor=colors[row[1].value])
widths = [8, 20, 14, 24, 40, 35, 12, 14, 12, 14, 12, 14, 12, 14, 16, 13, 13, 13, 12, 17, 48]
for i, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width
ws.freeze_panes = "G2"
if results:
    table = Table(displayName="ChongqingSafetyChoices", ref=f"A1:U{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(table)

note = wb["学科标注说明"]
note.append(("重庆保底表", "按近年同校同名专业位次中位数分层。准保底仍可能受招生计划和专业热度影响；真正保底应至少保留公办稳妥保底或本人能接受的民办最终兜底。"))
wb.save(BOOK)
print(BOOK)
from collections import Counter
print("rows:", len(results), Counter(r["category"] for r in results))
