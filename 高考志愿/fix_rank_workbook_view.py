import os

from openpyxl import load_workbook
from openpyxl.worksheet.views import Pane, Selection


path = os.environ["TARGET_XLSX"]
workbook = load_workbook(path)
sheet = workbook.worksheets[4]

for row_number in range(1, sheet.max_row + 1):
    sheet.row_dimensions[row_number].hidden = False

sheet.sheet_view.pane = Pane(
    ySplit=1,
    topLeftCell="A2",
    activePane="bottomLeft",
    state="frozen",
)
sheet.sheet_view.selection = [
    Selection(pane="bottomLeft", activeCell="A2", sqref="A2")
]
sheet.sheet_view.topLeftCell = "A1"
sheet.auto_filter.ref = f"A1:G{sheet.max_row}"

workbook.save(path)
