import re
import statistics
import time
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).parent
OUT = ROOT / "全国本科_重庆物理类位次20000至40000_学校专业详表.xlsx"
LOW, HIGH = 20000, 40000
RANK_URLS = {
    2022: "https://www.dxsbb.com/news/117895.html",
    2023: "https://www.dxsbb.com/news/136970.html",
    2024: "https://www.dxsbb.com/news/146467.html",
    2025: "https://www.dxsbb.com/news/148772.html",
}
PROVINCES = {
    "11": "北京", "12": "天津", "13": "河北", "14": "山西", "15": "内蒙古",
    "21": "辽宁", "22": "吉林", "23": "黑龙江", "31": "上海", "32": "江苏",
    "33": "浙江", "34": "安徽", "35": "福建", "36": "江西", "37": "山东",
    "41": "河南", "42": "湖北", "43": "湖南", "44": "广东", "45": "广西",
    "46": "海南", "50": "重庆", "51": "四川", "52": "贵州", "53": "云南",
    "54": "西藏", "61": "陕西", "62": "甘肃", "63": "青海", "64": "宁夏", "65": "新疆",
}
DOUBLE_FIRST = {
    "北京交通大学", "北京科技大学", "北京化工大学", "北京邮电大学", "北京林业大学",
    "北京中医药大学", "中国农业大学", "中央财经大学", "对外经济贸易大学",
    "华北电力大学", "北京工业大学", "天津医科大学", "河北工业大学", "太原理工大学",
    "内蒙古大学", "辽宁大学", "大连海事大学", "延边大学", "东北师范大学",
    "哈尔滨工程大学", "东北林业大学", "东北农业大学", "上海大学", "东华大学",
    "华东理工大学", "苏州大学", "南京航空航天大学", "南京理工大学", "中国矿业大学",
    "河海大学", "江南大学", "南京农业大学", "中国药科大学", "南京师范大学",
    "安徽大学", "合肥工业大学", "福州大学", "南昌大学", "中国石油大学(华东)",
    "郑州大学", "华中农业大学", "华中师范大学", "中南财经政法大学", "武汉理工大学",
    "湖南师范大学", "暨南大学", "华南师范大学", "广西大学", "海南大学",
    "西南大学", "西南交通大学", "西南财经大学", "四川农业大学", "贵州大学",
    "云南大学", "西北大学", "西安电子科技大学", "陕西师范大学", "长安大学",
    "西北农林科技大学", "兰州大学", "青海大学", "宁夏大学", "新疆大学", "石河子大学",
}
INDUSTRY = {
    "燕山大学", "东北电力大学", "南京工业大学", "南京信息工程大学", "浙江工业大学",
    "杭州电子科技大学", "广东工业大学", "长沙理工大学", "西安理工大学",
    "重庆邮电大学", "重庆交通大学", "重庆医科大学", "西南政法大学",
    "上海电力大学", "大连交通大学", "华东交通大学", "兰州交通大学", "石家庄铁道大学",
}
PRIVATE_MARKERS = ("独立学院", "职业技术大学", "城市学院", "科技学院", "财经学院", "工程学院")


def rank_map(year):
    last = None
    for attempt in range(5):
        try:
            table = pd.read_html(RANK_URLS[year])[0]
            break
        except Exception as exc:
            last = exc
            time.sleep(attempt + 1)
    else:
        raise last
    table.columns = ["score", "count", "rank"]
    out = {}
    for _, row in table.iterrows():
        m = re.search(r"\d+", str(row["score"]))
        if m:
            try:
                out[int(m.group())] = int(float(row["rank"]))
            except (TypeError, ValueError):
                pass
    return out


def plan_type(school, major):
    text = school + major
    if "非西藏生定藏就业" in text:
        return "定藏就业"
    if "地方专项" in text:
        return "地方专项"
    if "国家专项" in text:
        return "国家专项"
    if "高校专项" in text:
        return "高校专项"
    if "中外合作" in text:
        return "中外合作"
    if "民族班" in text:
        return "民族班"
    if "预科班" in text:
        return "预科班"
    return "普通计划"


def clean_school(name):
    return re.sub(r"\((地方专项|国家专项|高校专项|非西藏生定藏就业|中外合作|民族班|预科班).*?\)", "", name).strip()


def clean_major(name):
    text = re.sub(r"\((地方专项|国家专项计划?|高校专项|非西藏生定藏就业|中外合作办学).*?\)", "", name)
    return re.sub(r"\(师范类\)|\(师范\)", "", text).strip()


def tier(school):
    if school in DOUBLE_FIRST:
        return "双一流/原211"
    if school in INDUSTRY:
        return "行业强校"
    if any(x in school for x in PRIVATE_MARKERS):
        return "民办/独立学院（核验）"
    return "普通公办/性质待核验"


def subject_check(major):
    if "思想政治" in major:
        return "物化地通常不适配"
    return "物化地初筛可报；以2026计划为准"


def direction(major):
    rules = [
        (("计算机", "软件", "网络空间", "信息安全", "人工智能", "数据科学"), "计算机与信息"),
        (("电气", "自动化", "电子信息", "通信", "测控"), "电子电气与自动化"),
        (("车辆", "轨道", "交通", "铁道"), "车辆与交通"),
        (("储能", "能源", "动力"), "能源动力"),
        (("食品",), "食品科学与安全"), (("临床", "口腔", "医学", "药学"), "医药卫生"),
        (("水利", "水文"), "水利工程"), (("机械", "智能制造"), "机械制造"),
        (("材料", "化学", "化工"), "材料化工"), (("师范", "教育"), "师范教育"),
        (("经济", "金融", "会计", "工商管理"), "经济管理"),
        (("环境", "生态"), "环境生态"), (("法学",), "法学"),
    ]
    for words, label in rules:
        if any(word in major for word in words):
            return label
    return "其他"


ranks = {year: rank_map(year) for year in range(2022, 2026)}
year_rows = {}
for year in range(2022, 2026):
    data = []
    for raw in (ROOT / "data" / f"{year}.txt").read_text(encoding="utf-8").splitlines():
        line = " ".join(raw.split())
        m = re.match(r"^(\d{4})\s+(.+?)\s+([0-9A-Z]{3})\s+(.+?)\s+(\d{3})(?:\s|$)", line)
        if not m:
            continue
        code, school, major_code, major, score = m.groups()
        score = int(score)
        rank = ranks[year].get(score)
        if rank:
            data.append({
                "code": code, "school": school, "school_key": clean_school(school),
                "major_code": major_code, "major": major, "major_key": clean_major(major),
                "type": plan_type(school, major), "score": score, "rank": rank,
            })
    year_rows[year] = data

# 补入专项提取的定藏就业数据，禁止与普通计划混用。
tibet_file = ROOT / "重庆物理类_非西藏生定藏就业_筛选表.xlsx"
if tibet_file.exists():
    tbook = load_workbook(tibet_file, data_only=True)
    for record in tbook["2022-2025全部记录"].iter_rows(min_row=2, values_only=True):
        year, code, school, major_code, major, score, rank = record
        if year in year_rows and rank:
            year_rows[year].append({
                "code": str(code), "school": str(school), "school_key": clean_school(str(school)),
                "major_code": str(major_code), "major": str(major), "major_key": clean_major(str(major)),
                "type": "定藏就业", "score": int(score), "rank": int(rank),
            })

index = {}
for year, rows in year_rows.items():
    for row in rows:
        index.setdefault((year, row["school_key"], row["major_key"], row["type"]), []).append(row)

results, seen = [], set()
for current in year_rows[2025]:
    key = (current["school_key"], current["major_key"], current["type"])
    if key in seen:
        continue
    seen.add(key)
    history = {}
    for year in range(2022, 2026):
        matches = index.get((year, *key), [])
        if matches:
            match = sorted(matches, key=lambda r: len(r["major"]))[0]
            history[year] = (match["score"], match["rank"])
        else:
            history[year] = (None, None)
    valid = [rank for _, rank in history.values() if rank]
    if not valid:
        continue
    median = round(statistics.median(valid))
    # 主口径为历年中位数；仅一年数据时，以2025位次入选并标低置信度。
    if not (LOW <= median <= HIGH):
        continue
    volatility = max(valid) - min(valid)
    confidence = "高（4年）" if len(valid) == 4 else (
        "中（3年）" if len(valid) == 3 else ("较低（2年）" if len(valid) == 2 else "低（仅2025）")
    )
    results.append({
        **current, "history": history, "median": median, "best": min(valid), "worst": max(valid),
        "volatility": volatility, "years": len(valid), "confidence": confidence,
    })

type_order = {"普通计划": 1, "中外合作": 2, "地方专项": 3, "国家专项": 4,
              "高校专项": 5, "定藏就业": 6, "民族班": 7, "预科班": 8}
results.sort(key=lambda r: (r["median"], type_order.get(r["type"], 9), r["school_key"], r["major"]))

wb = Workbook()
ws = wb.active
ws.title = "全国学校专业详表"
headers = [
    "序号", "省份", "院校层次", "院校代码(2025)", "院校", "专业代码(2025)", "专业",
    "专业方向", "招生类型", "物化地适配初筛",
    "2022最低分", "2022最低位次", "2023最低分", "2023最低位次",
    "2024最低分", "2024最低位次", "2025最低分", "2025最低位次",
    "近4年位次中位数", "最好位次", "最差位次", "位次波动幅度",
    "有效年份", "数据置信度", "相对26954名差值", "资格/风险提示",
]
ws.append(headers)
for i, row in enumerate(results, 1):
    h = row["history"]
    risk = {
        "普通计划": "普通本科批；核验2026招生计划",
        "中外合作": "重点核验学费、培养模式和出国要求",
        "地方专项": "仅限通过地方专项资格审核",
        "国家专项": "仅限通过国家专项资格审核",
        "高校专项": "仅限通过高校专项资格及该校审核",
        "定藏就业": "须接受赴藏就业、服务年限及违约条款",
        "民族班": "须符合民族班报考资格",
        "预科班": "须符合预科班报考资格",
    }.get(row["type"], "核验招生资格")
    ws.append([
        i, PROVINCES.get(row["code"][:2], "待核验"), tier(row["school_key"]),
        row["code"], row["school_key"], row["major_code"], row["major"],
        direction(row["major"]), row["type"], subject_check(row["major"]),
        h[2022][0], h[2022][1], h[2023][0], h[2023][1],
        h[2024][0], h[2024][1], h[2025][0], h[2025][1],
        row["median"], row["best"], row["worst"], row["volatility"],
        row["years"], row["confidence"], row["median"] - 26954, risk,
    ])

for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    if row[8].value != "普通计划":
        row[8].fill = PatternFill("solid", fgColor="FFF2CC")
widths = [8, 9, 22, 15, 28, 16, 42, 19, 14, 28, 12, 14, 12, 14, 12, 14, 12, 14, 17, 13, 13, 15, 12, 15, 17, 48]
for i, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width
ws.freeze_panes = "K2"
table = Table(displayName="NationalRank24to30", ref=f"A1:Z{ws.max_row}")
table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(table)

summary = wb.create_sheet("分类统计")
summary.append(["分类维度", "类别", "数量"])
for label, counter in (
    ("招生类型", Counter(r["type"] for r in results)),
    ("省份", Counter(PROVINCES.get(r["code"][:2], "待核验") for r in results)),
    ("院校层次", Counter(tier(r["school_key"]) for r in results)),
    ("专业方向", Counter(direction(r["major"]) for r in results)),
):
    for category, count in counter.most_common():
        summary.append([label, category, count])
for cell in summary[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="548235")
summary.column_dimensions["A"].width = 18
summary.column_dimensions["B"].width = 35
summary.column_dimensions["C"].width = 12

guide = wb.create_sheet("口径与使用说明")
notes = [
    ("项目", "说明"),
    ("筛选范围", "重庆物理类本科批，近4年有效最低位次中位数20000—40000名；覆盖全国院校，不限制数量。"),
    ("考生", "2026重庆考生，物化地，566分，26954名。"),
    ("位次口径", "优先使用同院校、同专业、同招生类型的2022—2025数据；特殊类型绝不以普通计划数据替代。"),
    ("数据置信度", "4年为高、3年为中、2年为较低、仅2025为低。专业改名或新增会减少有效年份。"),
    ("院校层次", "双一流/原211及行业强校为辅助标签；其余院校性质和层次须以教育部及学校官网核验。"),
    ("选科", "物化地仅作初筛。2026实际选科、体检、语种、单科成绩和校区要求以正式计划为准。"),
    ("排序", "主表按位次中位数从小到大，即往年门槛由高到低。可利用筛选器按省份、层次、专业方向和招生类型筛选。"),
]
for row in notes:
    guide.append(row)
guide.column_dimensions["A"].width = 20
guide.column_dimensions["B"].width = 118
for cell in guide[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="7030A0")
for row in guide.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

src = wb.create_sheet("数据来源")
src.append(["年份", "专业投档表", "一分一段表"])
pdfs = {
    2022: "https://cdn.gaokzx.com/zixunzhan/202403/a8849ca2-9596-4d78-8eed-a9c3ac89df49.pdf",
    2023: "https://cdn.gaokzx.com/zixunzhan/202403/88fefb53-5c97-4a16-bfa2-565abcbe37dd.pdf",
    2024: "https://cdn.zizzs.com/1721564943406%E7%89%A9%E7%90%86.pdf",
    2025: "https://cdn.gaokzx.com/zixunzhan/1753323564725%E7%89%A9%E7%90%86.pdf",
}
for year in range(2022, 2026):
    src.append([year, pdfs[year], RANK_URLS[year]])
for row in range(2, src.max_row + 1):
    for col in (2, 3):
        src.cell(row, col).hyperlink = src.cell(row, col).value
        src.cell(row, col).style = "Hyperlink"
for cell in src[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="44546A")
src.column_dimensions["A"].width = 10
src.column_dimensions["B"].width = 100
src.column_dimensions["C"].width = 55

wb.save(OUT)
print(OUT)
print("rows:", len(results), "schools:", len({r["school_key"] for r in results}))
