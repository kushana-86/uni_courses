import re
import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).parent
BOOK = ROOT / "2026重庆高考_提前批AB段加本科批_整体填报方案.xlsx"
RANK_URL = "https://www.dxsbb.com/news/148772.html"
TARGET = 26954


def rank_map():
    last = None
    for attempt in range(5):
        try:
            table = pd.read_html(RANK_URL)[0]
            break
        except Exception as exc:
            last = exc
            time.sleep(attempt + 1)
    else:
        raise last
    table.columns = ["score", "count", "rank"]
    out = {}
    for _, row in table.iterrows():
        m = re.search(r"\d+", str(row["score"]))
        if m:
            try:
                out[int(m.group())] = int(float(row["rank"]))
            except (TypeError, ValueError):
                pass
    return out


def strength(school, major=""):
    s = school or ""
    m = major or ""
    explicit = {
        "江南大学": "食品科学与工程、轻工技术与工程",
        "中国农业大学": "农业工程、食品科学、作物学、畜牧学",
        "中国海洋大学": "海洋科学、水产、食品科学",
        "华中农业大学": "作物学、园艺学、畜牧学、食品科学",
        "东北农业大学": "农业工程、食品科学、畜牧学",
        "南京农业大学": "作物学、农业资源与环境、农林经济管理",
        "河南工业大学": "粮食工程、食品科学与工程",
        "武汉轻工大学": "食品科学、粮油工程",
        "华北电力大学": "电气工程、动力工程",
        "东北电力大学": "电气工程、动力工程",
        "长沙理工大学": "交通运输工程、土木工程、电气工程",
        "中国石油大学": "石油与天然气工程、地质资源、化工",
        "南京航空航天大学": "航空宇航、力学、控制科学",
        "西南交通大学": "交通运输工程、土木工程、电气工程",
        "北京交通大学": "交通运输工程、信息与通信工程、系统科学",
        "中南大学": "交通运输工程、冶金工程、矿业工程、临床医学",
        "长安大学": "交通运输工程、道路桥梁、车辆工程、地质工程",
        "兰州交通大学": "交通运输工程、土木工程、自动控制",
        "华东交通大学": "交通运输工程、控制科学、土木工程",
        "石家庄铁道大学": "土木工程、交通运输工程、机械工程",
        "大连交通大学": "机械工程、交通运输工程、材料工程",
        "武汉理工大学": "材料科学、交通运输工程、机械工程",
        "吉林大学": "车辆工程、机械工程、化学、法学",
        "合肥工业大学": "车辆工程、机械工程、管理科学与工程",
        "燕山大学": "机械工程、材料科学、控制科学",
        "重庆理工大学": "车辆工程、机械工程、材料工程",
        "湖北汽车工业学院": "车辆工程、机械工程、汽车产业应用方向",
        "重庆邮电大学": "信息与通信工程、计算机、控制科学",
        "重庆交通大学": "交通运输工程、土木工程、水利工程",
        "西南大学": "教育学、农业科学、生物学、食品科学",
        "重庆大学": "机械工程、电气工程、建筑学、土木工程",
        "西南政法大学": "法学",
        "上海电力大学": "电气工程、动力工程",
        "沈阳药科大学": "药学、中药学",
        "南京信息工程大学": "大气科学、计算机、环境科学",
    }
    for key, value in explicit.items():
        if key in s:
            return value
    rules = [
        ("医科大学", "临床医学、基础医学、公共卫生"),
        ("中医药大学", "中医学、中药学、针灸推拿"),
        ("农业大学", "农业科学、食品科学、生物学"),
        ("农林大学", "林学、农业科学、食品科学"),
        ("财经大学", "应用经济学、工商管理、统计学"),
        ("工商大学", "工商管理、应用经济学、食品科学"),
        ("师范大学", "教育学、中国语言文学、数学及基础学科"),
        ("外国语大学", "外国语言文学"),
        ("政法大学", "法学"),
        ("石油大学", "石油与天然气工程、地质资源、化工"),
        ("电力大学", "电气工程、动力工程"),
        ("交通大学", "交通运输工程、土木工程、机械工程"),
        ("海洋大学", "海洋科学、水产、食品科学"),
        ("航空航天大学", "航空宇航、力学、控制科学"),
        ("邮电大学", "信息与通信工程、计算机"),
        ("科技大学", "工科与材料、机械、信息类（按具体学校核验）"),
    ]
    for key, value in rules:
        if key in s:
            return value
    if "食品" in m:
        return "本校食品科学/生物与化工方向（层次须核验）"
    if "储能" in m:
        return "本校能源动力/电气/材料方向（层次须核验）"
    if any(x in m for x in ("计算机", "软件", "电子信息", "通信")):
        return "信息类相关方向（是否优势学科须核验）"
    if any(x in m for x in ("机械", "自动化", "电气")):
        return "工科相关方向（是否优势学科须核验）"
    return "见院校学科评估及一流学科名单"


ranks = rank_map()
special = []
for raw in (ROOT / "data" / "2025.txt").read_text(encoding="utf-8").splitlines():
    line = " ".join(raw.split())
    if not any(k in line for k in (
        "储能科学与工程", "食品质量与安全", "食品科学与工程", "食品安全与检测",
        "车辆工程", "轨道交通", "交通设备与控制", "铁道工程",
    )):
        continue
    m = re.match(r"^(\d{4})\s+(.+?)\s+([0-9A-Z]{3})\s+(.+?)\s+(\d{3})(?:\s|$)", line)
    if not m:
        continue
    code, school, major_code, major, score = m.groups()
    score = int(score)
    rank = ranks.get(score)
    if not rank:
        continue
    special.append({
        "batch": "普通本科批", "code": code, "school": school, "major_code": major_code,
        "major": major, "score": score, "rank": rank,
    })

for raw in (ROOT / "data" / "2025_B_layout.txt").read_text(encoding="utf-8").splitlines():
    line = " ".join(raw.split())
    if not any(k in line for k in (
        "储能科学与工程", "食品质量与安全", "食品科学与工程", "食品安全与检测",
        "车辆工程", "轨道交通", "交通设备与控制", "铁道工程",
    )):
        continue
    m = re.match(r"^(\d{4})\s*(.+?)\s+([0-9A-Z]{3})\s*(.+?)\s*(\d{3})(?:\s|$)", line)
    if not m:
        continue
    code, school, major_code, major, score = m.groups()
    score = int(score)
    rank = ranks.get(score)
    if not rank:
        continue
    batch = "B段-高校专项" if "高校专项" in school + major else "B段-国家专项"
    special.append({
        "batch": batch, "code": code, "school": school, "major_code": major_code,
        "major": major, "score": score, "rank": rank,
    })

# 仅用于排序；不再限制院校或专业数量。
def special_value(row):
    text = row["school"] + row["major"]
    value = 0
    for key in ("江南大学", "中国农业大学", "中国海洋大学", "华中农业大学",
                "东北农业大学", "河南工业大学", "武汉轻工大学", "华北电力大学",
                "东北电力大学", "长沙理工大学", "中国石油大学", "南京航空航天大学"):
        if key in text:
            value += 10
    if "储能" in row["major"]:
        value += 4
    if "食品质量与安全" in row["major"]:
        value += 3
    return value


special.sort(key=lambda r: (r["rank"], -special_value(r), r["school"], r["major"]))
chosen, seen = [], set()
for row in special:
    key = (row["school"], row["major"], row["batch"])
    if key in seen:
        continue
    seen.add(key)
    chosen.append(row)


def direction(major):
    if "储能" in major:
        return "储能与能源"
    if any(x in major for x in ("食品质量与安全", "食品安全与检测")):
        return "食品安全"
    if "食品科学与工程" in major:
        return "食品科学"
    if any(x in major for x in ("轨道交通", "铁道工程", "交通设备与控制")):
        return "轨道交通"
    if "车辆工程" in major:
        return "车辆工程"
    return "相关方向"

wb = load_workbook(BOOK)
sheet_specs = {
    "A段条件式备选": (3, None),
    "B段建议50项": (4, 6),
    "普通本科批100项": (4, 7),
    "定藏就业条件备选": (3, 5),
}
for name, (school_col, major_col) in sheet_specs.items():
    if name not in wb.sheetnames:
        continue
    ws = wb[name]
    existing = [c.column for c in ws[1] if c.value == "院校强势学科/特色方向"]
    if existing:
        new_col = existing[0]
        for duplicate in reversed(existing[1:]):
            ws.delete_cols(duplicate, 1)
    else:
        new_col = ws.max_column + 1
    ws.cell(1, new_col, "院校强势学科/特色方向")
    ws.cell(1, new_col).font = Font(bold=True, color="FFFFFF")
    ws.cell(1, new_col).fill = PatternFill("solid", fgColor="1F4E78")
    ws.cell(1, new_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(2, ws.max_row + 1):
        school = ws.cell(r, school_col).value or ""
        major = ws.cell(r, major_col).value if major_col else ""
        ws.cell(r, new_col, strength(str(school), str(major or "")))
        ws.cell(r, new_col).alignment = Alignment(vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(new_col)].width = 42

for old_name in ("储能与食品安全专项", "重点专业完整汇总"):
    if old_name in wb.sheetnames:
        del wb[old_name]
ws = wb.create_sheet("重点专业完整汇总", 1)
headers = ["序号", "专业方向", "批次/类型", "院校代码", "院校", "专业代码", "专业",
           "2025最低分", "2025最低位次", "与本人位次差", "梯度",
           "院校强势学科/特色方向", "资格、学费及填报提示"]
ws.append(headers)
for i, row in enumerate(chosen, 1):
    diff = row["rank"] - TARGET
    band = "冲" if diff < -3000 else ("稳" if diff <= 4000 else "保")
    qualification = "须通过该校高校专项审核" if "高校专项" in row["batch"] else (
        "须通过国家专项资格审核" if "国家专项" in row["batch"] else "普通本科批可直接按计划填报"
    )
    cooperation = "；中外合作项目须重点核验学费和培养模式" if "中外合作" in row["school"] + row["major"] else ""
    ws.append([i, direction(row["major"]), row["batch"], row["code"], row["school"], row["major_code"], row["major"],
               row["score"], row["rank"], diff, band, strength(row["school"], row["major"]),
               qualification + cooperation + "；选科和体检以2026计划为准"])
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="548235")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    row[10].fill = PatternFill("solid", fgColor={"冲": "F4CCCC", "稳": "FFF2CC", "保": "D9EAD3"}[row[10].value])
widths = [8, 14, 18, 12, 28, 12, 40, 13, 15, 16, 9, 42, 52]
for i, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width
ws.freeze_panes = "H2"
table = Table(displayName="FocusMajorChoices", ref=f"A1:M{ws.max_row}")
table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
ws.add_table(table)

if "学科标注说明" in wb.sheetnames:
    del wb["学科标注说明"]
note = wb.create_sheet("学科标注说明")
notes = [
    ("项目", "说明"),
    ("强势学科口径", "用于志愿比较的院校传统优势/特色方向摘要，不等同于教育部最新学科评估名次，也不代表表中专业一定属于该校最强学科。"),
    ("核验建议", "最终决策前查看教育部第二轮双一流建设学科名单、学校官网学科介绍、国家级一流本科专业及培养方案。"),
    ("储能专业", "重点比较电气工程、动力工程、材料科学、化学工程等学科基础，以及实验平台和行业合作。"),
    ("食品安全", "重点比较食品科学与工程、轻工、农业、生物、化学和公共卫生基础；食品质量与安全与食品科学与工程培养方向不同。"),
    ("车辆工程", "重点比较机械工程、车辆工程学科基础，新能源汽车、智能网联、动力系统方向及汽车产业所在地。"),
    ("轨道交通", "重点比较交通运输工程、控制科学、电气工程、土木工程基础；区分信号控制、车辆、铁道工程和运营管理。"),
    ("数量口径", "专题表不限制院校或专业数量，保留2025原始投档表中可结构化识别的全部相关项目，包括专项及中外合作并明确提示。"),
]
for row in notes:
    note.append(row)
note.column_dimensions["A"].width = 22
note.column_dimensions["B"].width = 115
for cell in note[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="548235")
for row in note.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

wb.save(BOOK)
print(BOOK)
print("special rows:", len(chosen))
