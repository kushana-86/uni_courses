from pathlib import Path

import build_recommendations as base
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

segments = [(25000, 26000, 12), (26000, 27000, 13), (27000, 28000, 13), (28000, 29001, 12)]
chosen = []
used = set()
for low, high, count in segments:
    pool = [c for c in base.diverse if low <= c["median_rank"] < high]
    pool.sort(key=lambda c: (-c["score_desire"], -c["years"], c["volatility"]))
    for c in pool:
        key = (c["school"], c["major"])
        if key not in used:
            chosen.append(c)
            used.add(key)
        if sum(low <= x["median_rank"] < high for x in chosen) >= count:
            break
chosen.sort(key=lambda c: c["median_rank"])

wb = Workbook()
ws = wb.active
ws.title = "2.5万-2.9万大学专业"
headers = [
    "序号", "院校", "省份", "专业", "近4年位次中位数", "相对本人26954名",
    "2022最低分", "2022最低位次", "2023最低分", "2023最低位次",
    "2024最低分", "2024最低位次", "2025最低分", "2025最低位次",
    "位次波动范围", "有效年份", "简要判断",
]
ws.append(headers)
for i, c in enumerate(chosen, 1):
    h = c["history"]
    diff = c["median_rank"] - base.TARGET_RANK
    if diff < -1000:
        judgement = "略偏冲"
    elif diff <= 1500:
        judgement = "与本人位次接近"
    else:
        judgement = "略偏稳"
    ws.append([
        i, c["school"], base.PROVINCES.get(c["code"][:2], "待核验"), c["major"],
        c["median_rank"], diff,
        h[2022][0], h[2022][1], h[2023][0], h[2023][1],
        h[2024][0], h[2024][1], h[2025][0], h[2025][1],
        c["volatility"], c["years"], judgement,
    ])

table = Table(displayName="RankBandSummary", ref=f"A1:Q{ws.max_row}")
table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(table)
ws.freeze_panes = "G2"
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    color = {"略偏冲": "F4CCCC", "与本人位次接近": "FFF2CC", "略偏稳": "D9EAD3"}[row[16].value]
    row[16].fill = PatternFill("solid", fgColor=color)
widths = [8, 25, 9, 36, 17, 16, 12, 14, 12, 14, 12, 14, 12, 14, 14, 11, 18]
for i, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

note = wb.create_sheet("口径说明")
notes = [
    ("项目", "说明"),
    ("范围", "近4年有效投档位次的中位数在25000—29000名之间，共50个院校+专业组合。"),
    ("考生基准", "重庆物理类，物化地，566分，26954名。"),
    ("判断", "略偏冲：中位位次比本人高1000名以上；接近：差值-1000至+1500；略偏稳：比本人低1500名以上。"),
    ("提醒", "同一专业年度波动可能较大；最低分、代码、招生人数和选科要求须以2026重庆招生计划为准。"),
    ("缺失", "空白表示当年没有检出完全匹配的同校同名专业，不使用推测值填补。"),
]
for row in notes:
    note.append(row)
note.column_dimensions["A"].width = 18
note.column_dimensions["B"].width = 105
for cell in note[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
for row in note.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

out = Path(__file__).parent / "重庆物理类_位次2万5至2万9_大学专业汇总.xlsx"
wb.save(out)
print(out)
print("rows:", len(chosen))
