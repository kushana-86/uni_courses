import re
import statistics
import time
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).parent
INPUT = min((ROOT / "新建文件夹").glob("*.xlsx"), key=lambda p: p.stat().st_size)
OUTPUT = INPUT.parent / "本科批_四种排序方案.xlsx"
TARGET_RANK = 26954
RANK_URLS = {
    2022: "https://www.dxsbb.com/news/117895.html",
    2023: "https://www.dxsbb.com/news/136970.html",
    2024: "https://www.dxsbb.com/news/146467.html",
    2025: "https://www.dxsbb.com/news/148772.html",
}

SCHOOL_SCORE = {
    "湖南大学": 98, "河海大学": 95, "西南大学": 93, "西北农林科技大学": 92,
    "江南大学": 91, "合肥工业大学": 90, "陕西师范大学": 90, "北京林业大学": 89,
    "湘潭大学": 87, "南京信息工程大学": 86, "西南石油大学": 85,
    "长沙理工大学": 83, "重庆邮电大学": 83, "中南民族大学": 82,
    "东北电力大学": 81, "兰州交通大学": 80, "石家庄铁道大学": 79,
    "西南政法大学": 79, "重庆医科大学": 79, "华东交通大学": 78,
    "北方工业大学": 78, "三峡大学": 78, "中北大学": 77,
    "重庆交通大学": 77, "重庆理工大学": 76, "华北水利水电大学": 76,
    "上海工程技术大学": 75, "中国民用航空飞行学院": 75,
    "西华大学": 73, "重庆师范大学": 73, "重庆科技大学": 70,
    "云南警官学院": 69, "贵州警察学院": 68, "重庆文理学院": 67,
    "长江师范学院": 65, "广东石油化工学院": 64, "辽宁科技大学": 64,
}
SCHOOL_STRENGTH = {
    "河海大学": ("水利", "水文", "储能", "电气"),
    "江南大学": ("食品", "轻工", "生物"),
    "西南大学": ("食品", "心理", "化学", "自动化", "教育"),
    "湖南大学": ("车辆", "机械", "电气", "自动化"),
    "合肥工业大学": ("车辆", "机械", "管理", "电气"),
    "西北农林科技大学": ("水利", "林学", "农业", "食品"),
    "南京信息工程大学": ("大气", "气象", "计算机"),
    "西南石油大学": ("储能", "石油", "新能源", "机械"),
    "长沙理工大学": ("储能", "电气", "交通", "水利"),
    "重庆邮电大学": ("网络", "计算机", "物联网", "自动化", "测控", "车辆"),
    "东北电力大学": ("储能", "电气", "能源"),
    "兰州交通大学": ("轨道", "车辆", "铁道", "交通", "电气"),
    "石家庄铁道大学": ("轨道", "铁道", "交通", "土木"),
    "华东交通大学": ("轨道", "交通", "铁道", "车辆", "电气"),
    "三峡大学": ("水利", "水文", "电网", "电气"),
    "中北大学": ("车辆", "装甲", "弹药", "兵器"),
    "重庆交通大学": ("交通", "车辆", "水利", "土木"),
    "重庆理工大学": ("车辆", "装甲", "机械", "自动化", "会计"),
    "华北水利水电大学": ("水利", "水文"),
    "中国民用航空飞行学院": ("气象", "交通", "航空", "计算机"),
    "重庆科技大学": ("石油", "自动化", "储能", "冶金"),
}


def rank_map(year):
    last = None
    for attempt in range(5):
        try:
            table = pd.read_html(RANK_URLS[year])[0]
            break
        except Exception as exc:
            last = exc
            time.sleep(attempt + 1)
    else:
        raise last
    table.columns = ["score", "count", "rank"]
    out = {}
    for _, row in table.iterrows():
        m = re.search(r"\d+", str(row["score"]))
        if m:
            try:
                out[int(m.group())] = int(float(row["rank"]))
            except (TypeError, ValueError):
                pass
    return out


def base_school(text):
    return re.sub(r"\((地方专项|国家专项|高校专项|非西藏生定藏就业|中外合作).*?\)", "", text or "").strip()


def base_major(text):
    text = text or ""
    text = re.sub(r"\(地方专项\)", "", text)
    text = re.sub(r"\(非西藏生定藏就业\)", "", text)
    text = re.sub(r"\(师范类\)|\(师范\)", "", text)
    return text.strip()


def plan_type(school, major):
    text = (school or "") + (major or "")
    if "非西藏生定藏就业" in text:
        return "定藏就业"
    if "地方专项" in text:
        return "地方专项"
    if "国家专项" in text:
        return "国家专项"
    return "普通计划"


def major_score(major):
    rules = [
        (("电气工程", "智能电网"), 95), (("网络空间安全", "信息安全"), 94),
        (("计算机科学", "计算机类"), 93), (("自动化",), 91),
        (("物联网工程",), 90), (("轨道交通信号",), 90),
        (("应用气象", "大气科学"), 89), (("水利水电",), 89),
        (("交通运输",), 87), (("智能车辆", "新能源汽车"), 87),
        (("储能科学",), 86), (("车辆工程",), 85),
        (("测控技术",), 84), (("水文与水资源", "智慧水利"), 84),
        (("石油工程",), 83), (("食品科学", "食品质量"), 82),
        (("铁道工程", "交通工程"), 81), (("机械设计",), 80),
        (("装甲车辆",), 79), (("弹药工程",), 78),
        (("物理学", "化学"), 77), (("会计学",), 76),
        (("心理学",), 74), (("英语", "地理科学", "生物科学"), 73),
        (("林学",), 70),
    ]
    for words, score in rules:
        if any(word in major for word in words):
            return score
    return 72


rank_maps = {year: rank_map(year) for year in range(2022, 2026)}
admissions = {}
for year in range(2022, 2026):
    rows = []
    for raw in (ROOT / "data" / f"{year}.txt").read_text(encoding="utf-8").splitlines():
        line = " ".join(raw.split())
        m = re.match(r"^(\d{4})\s+(.+?)\s+([0-9A-Z]{3})\s+(.+?)\s+(\d{3})(?:\s|$)", line)
        if not m:
            continue
        code, school, major_code, major, score = m.groups()
        score = int(score)
        rank = rank_maps[year].get(score)
        if rank:
            rows.append({
                "code": code, "school": school, "school_key": base_school(school),
                "major_code": major_code, "major": major, "major_key": base_major(major),
                "type": plan_type(school, major), "score": score, "rank": rank,
            })
    admissions[year] = rows

# 旧年度PDF中部分“非西藏生定藏就业”行存在院校名与专业代码粘连，
# 使用已经专项提取并核验的长表补齐，避免回退匹配到普通计划。
tibet_file = ROOT / "重庆物理类_非西藏生定藏就业_筛选表.xlsx"
if tibet_file.exists():
    tibet_book = load_workbook(tibet_file, data_only=True)
    if "2022-2025全部记录" in tibet_book.sheetnames:
        for record in tibet_book["2022-2025全部记录"].iter_rows(min_row=2, values_only=True):
            year, code, school, major_code, major, score, rank = record
            if year in admissions and school and major and rank:
                admissions[year].append({
                    "code": str(code), "school": str(school), "school_key": base_school(str(school)),
                    "major_code": str(major_code), "major": str(major), "major_key": base_major(str(major)),
                    "type": "定藏就业", "score": int(score), "rank": int(rank),
                })

index = {}
for year, rows in admissions.items():
    for row in rows:
        index.setdefault((year, row["school_key"], row["major_key"]), []).append(row)

source_book = load_workbook(INPUT, data_only=True)
source = source_book.active
items = []
for row in source.iter_rows(min_row=4, values_only=True):
    if not row[2] or not row[5]:
        continue
    school, major = str(row[2]).strip(), str(row[5]).strip()
    ptype = plan_type(school, major)
    history = {}
    for year in range(2022, 2026):
        matches = index.get((year, base_school(school), base_major(major)), [])
        if matches:
            typed = [m for m in matches if m["type"] == ptype]
            # 特殊招生类型不得用普通计划数据兜底；两者录取条件和分数口径不同。
            if ptype != "普通计划" and not typed:
                history[year] = (None, None)
            else:
                match = sorted(typed or matches, key=lambda x: (x["type"] != ptype, len(x["major"])))[0]
                history[year] = (match["score"], match["rank"])
        else:
            history[year] = (None, None)
    valid_ranks = [rank for _, rank in history.values() if rank]
    median_rank = round(statistics.median(valid_ranks)) if valid_ranks else None
    s_score = SCHOOL_SCORE.get(base_school(school), 68)
    p_score = major_score(major)
    synergy = 8 if any(k in major for k in SCHOOL_STRENGTH.get(base_school(school), ())) else 0
    # 录取匹配分：位次越接近本人越高；过度冲刺或过度保守都会降低综合排序。
    if median_rank:
        distance = abs(median_rank - TARGET_RANK)
        fit = max(35, 100 - distance / 650)
    else:
        fit = 45
    condition_penalty = 0
    warning = "普通本科批"
    if ptype == "地方专项":
        condition_penalty = 4
        warning = "仅限已通过地方专项资格"
    elif ptype == "定藏就业":
        condition_penalty = 8
        warning = "须接受赴藏定向就业、服务年限及违约条款"
    overall = round(s_score * 0.42 + min(100, p_score + synergy) * 0.38 + fit * 0.20 - condition_penalty, 2)
    items.append({
        "original_no": row[0], "school_code": row[1], "school": school, "nature": row[3],
        "major_code": row[4], "major": major, "remark": row[6], "type": ptype,
        "school_score": s_score, "major_score": p_score, "synergy": synergy,
        "overall": overall, "history": history, "median_rank": median_rank,
        "fit": round(fit, 1), "warning": warning,
    })

headers = [
    "新序号", "原志愿编号", "院校名称", "办学性质", "专业名称", "招生类型",
    "学校评分", "专业评分", "学科匹配加分", "综合评分",
    "2022最低位次", "2023最低位次", "2024最低位次", "2025最低位次",
    "近4年位次中位数", "与本人位次差", "资格/风险提示", "专业备注",
]


def write_sheet(wb, name, sorted_items, explanation):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for i, item in enumerate(sorted_items, 1):
        h = item["history"]
        median = item["median_rank"]
        ws.append([
            i, item["original_no"], item["school"], item["nature"], item["major"], item["type"],
            item["school_score"], item["major_score"], item["synergy"], item["overall"],
            h[2022][1], h[2023][1], h[2024][1], h[2025][1], median,
            median - TARGET_RANK if median else None, item["warning"], item["remark"],
        ])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        if row[5].value != "普通计划":
            row[5].fill = PatternFill("solid", fgColor="FFF2CC")
    widths = [9, 13, 27, 11, 42, 13, 11, 11, 14, 12, 14, 14, 14, 14, 17, 16, 48, 48]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "G2"
    table = Table(displayName="T" + re.sub(r"\W", "", name), ref=f"A1:R{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddFooter.center.text = explanation


wb = Workbook()
wb.remove(wb.active)
raw = wb.create_sheet("原始名单")
for row in source.iter_rows(values_only=True):
    raw.append(list(row))

school_sorted = sorted(items, key=lambda x: (-x["school_score"], -x["synergy"], -x["major_score"], x["median_rank"] or 999999))
major_sorted = sorted(items, key=lambda x: (-(x["major_score"] + x["synergy"]), -x["school_score"], x["median_rank"] or 999999))
overall_sorted = sorted(items, key=lambda x: (-x["overall"], abs((x["median_rank"] or 999999) - TARGET_RANK)))
rank_sorted = sorted(items, key=lambda x: (x["median_rank"] is None, x["median_rank"] or 999999))

write_sheet(wb, "方案1_学校优先", school_sorted, "先看学校层次，再看专业与位次")
write_sheet(wb, "方案2_专业优先", major_sorted, "先看专业质量及学校学科匹配，再看学校层次")
write_sheet(wb, "方案3_综合排序", overall_sorted, "学校42%+专业及学科匹配38%+位次匹配20%-条件限制")
write_sheet(wb, "方案4_位次由高到低", rank_sorted, "按近4年最低位次中位数从高位到低位排列")

guide = wb.create_sheet("评分说明", 1)
notes = [
    ("项目", "说明"),
    ("考生基准", "重庆物理类，物化地，566分，26954名。"),
    ("学校评分", "综合考虑办学层次、双一流/行业影响力、学校整体声誉；属于辅助排序，不是官方排名。"),
    ("专业评分", "综合考虑培养门槛、就业适配、发展空间；同一专业在不同学校通过“学科匹配加分”体现差异。"),
    ("综合评分", "学校42%＋专业及学科匹配38%＋录取位次匹配20%，地方专项和定藏就业按资格限制适度扣分。"),
    ("位次排序", "按2022—2025可匹配年份的最低位次中位数从小到大；小位次代表往年录取门槛更高。"),
    ("数据缺失", "专业改名、新增或招生类型变化会造成空白；未用普通计划分数冒充专项或定向计划。"),
    ("重要提醒", "学校优先、专业优先没有绝对答案。定藏就业和地方专项必须先确认资格与履约意愿，再参与排序。"),
]
for row in notes:
    guide.append(row)
guide.column_dimensions["A"].width = 22
guide.column_dimensions["B"].width = 115
for cell in guide[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="7030A0")
for row in guide.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

wb.save(OUTPUT)
print("input:", INPUT)
print("output:", OUTPUT)
print("items:", len(items))
