import re
import statistics
import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).parent
BOOK = ROOT / "2026重庆高考_提前批AB段加本科批_整体填报方案.xlsx"
TARGET = 26954
RANK_URLS = {
    2022: "https://www.dxsbb.com/news/117895.html",
    2023: "https://www.dxsbb.com/news/136970.html",
    2024: "https://www.dxsbb.com/news/146467.html",
    2025: "https://www.dxsbb.com/news/148772.html",
}

DOUBLE_FIRST = {
    "北京交通大学", "北京科技大学", "北京化工大学", "北京邮电大学", "北京林业大学",
    "北京中医药大学", "中国传媒大学", "中央财经大学", "对外经济贸易大学",
    "华北电力大学", "中国矿业大学(北京)", "中国石油大学(北京)", "中国地质大学(北京)",
    "北京工业大学", "天津医科大学", "河北工业大学", "太原理工大学", "内蒙古大学",
    "辽宁大学", "大连海事大学", "延边大学", "东北师范大学", "哈尔滨工程大学",
    "东北林业大学", "东北农业大学", "上海大学", "东华大学", "华东理工大学",
    "苏州大学", "南京航空航天大学", "南京理工大学", "中国矿业大学", "河海大学",
    "江南大学", "南京农业大学", "中国药科大学", "南京师范大学", "安徽大学",
    "合肥工业大学", "福州大学", "南昌大学", "中国石油大学(华东)", "郑州大学",
    "华中农业大学", "华中师范大学", "中南财经政法大学", "武汉理工大学",
    "湖南师范大学", "暨南大学", "华南师范大学", "广西大学", "海南大学",
    "西南大学", "西南交通大学", "西南财经大学", "四川农业大学", "贵州大学",
    "云南大学", "西北大学", "西安电子科技大学", "陕西师范大学", "长安大学",
    "西北农林科技大学", "兰州大学", "青海大学", "宁夏大学", "新疆大学", "石河子大学",
}
INDUSTRY_STRONG = {
    "燕山大学", "东北电力大学", "南京工业大学", "南京信息工程大学", "浙江工业大学",
    "杭州电子科技大学", "广东工业大学", "长沙理工大学", "西安理工大学", "西安建筑科技大学",
    "重庆邮电大学", "重庆交通大学", "重庆医科大学", "西南政法大学", "深圳技术大学",
    "上海电力大学", "大连交通大学", "华东交通大学", "兰州交通大学", "石家庄铁道大学",
}


def rank_map(year):
    last = None
    for attempt in range(5):
        try:
            table = pd.read_html(RANK_URLS[year])[0]
            break
        except Exception as exc:
            last = exc
            time.sleep(attempt + 1)
    else:
        raise last
    table.columns = ["score", "count", "rank"]
    out = {}
    for _, row in table.iterrows():
        m = re.search(r"\d+", str(row["score"]))
        if m:
            try:
                out[int(m.group())] = int(float(row["rank"]))
            except (ValueError, TypeError):
                pass
    return out


def clean_school(name):
    return re.sub(r"\((地方专项|国家专项|高校专项|中外合作|民族班|预科班).*?\)", "", name).strip()


def clean_major(name):
    name = re.sub(r"\(.*?\)", "", name)
    return re.sub(r"(类|试验班.*)$", "", name).strip()


rank_maps = {year: rank_map(year) for year in range(2022, 2026)}
year_rows = {}
for year in range(2022, 2026):
    rows = []
    for raw in (ROOT / "data" / f"{year}.txt").read_text(encoding="utf-8").splitlines():
        line = " ".join(raw.split())
        m = re.match(r"^(\d{4})\s+(.+?)\s+([0-9A-Z]{3})\s+(.+?)\s+(\d{3})(?:\s|$)", line)
        if not m:
            continue
        code, school, major_code, major, score = m.groups()
        if any(x in school for x in ("地方专项", "国家专项", "高校专项", "民族班", "预科班")):
            continue
        score = int(score)
        rank = rank_maps[year].get(score)
        if rank:
            rows.append({
                "code": code, "school": school, "school_key": clean_school(school),
                "major_code": major_code, "major": major, "major_key": clean_major(major),
                "score": score, "rank": rank,
            })
    year_rows[year] = rows

index = {}
for year, rows in year_rows.items():
    for row in rows:
        index.setdefault((year, row["school_key"], row["major_key"]), []).append(row)

results, seen = [], set()
for current in year_rows[2025]:
    school = current["school_key"]
    if school not in DOUBLE_FIRST and school not in INDUSTRY_STRONG:
        continue
    key = (school, current["major_key"])
    if key in seen:
        continue
    seen.add(key)
    history = {}
    for year in range(2022, 2026):
        matches = index.get((year, school, current["major_key"]), [])
        if matches:
            match = sorted(matches, key=lambda r: len(r["major"]))[0]
            history[year] = (match["score"], match["rank"])
        else:
            history[year] = (None, None)
    valid = [rank for _, rank in history.values() if rank]
    if len(valid) < 3:
        continue
    best, worst = min(valid), max(valid)
    volatility = worst - best
    median = round(statistics.median(valid))
    # 波动显著，且考生位次落在波动区间内或距离区间不超过5000名。
    distance = 0 if best <= TARGET <= worst else min(abs(TARGET - best), abs(TARGET - worst))
    if volatility < 6000 or distance > 5000:
        continue
    reachable_years = sum(rank >= TARGET for rank in valid)
    tier = "双一流/原211" if school in DOUBLE_FIRST else "行业强校"
    results.append({
        **current, "history": history, "best": best, "worst": worst, "volatility": volatility,
        "median": median, "years": len(valid), "reachable_years": reachable_years,
        "distance": distance, "tier": tier,
    })

results.sort(key=lambda r: (
    0 if r["best"] <= TARGET <= r["worst"] else 1,
    -r["reachable_years"], -r["volatility"], abs(r["median"] - TARGET),
))

wb = load_workbook(BOOK)
if "好学校波动机会" in wb.sheetnames:
    del wb["好学校波动机会"]
ws = wb.create_sheet("好学校波动机会", 2)
headers = [
    "序号", "院校层次", "院校", "专业", "2022最低分", "2022最低位次",
    "2023最低分", "2023最低位次", "2024最低分", "2024最低位次",
    "2025最低分", "2025最低位次", "4年位次中位数", "最好位次",
    "最差位次", "波动幅度", "本人可达年份数", "有效年份", "机会判断", "风险提示",
]
ws.append(headers)
for i, row in enumerate(results, 1):
    h = row["history"]
    if row["best"] <= TARGET <= row["worst"]:
        judgement = "波动区间覆盖本人位次"
    else:
        judgement = f"距历史波动区间约{row['distance']}名"
    ws.append([
        i, row["tier"], row["school_key"], row["major"],
        h[2022][0], h[2022][1], h[2023][0], h[2023][1],
        h[2024][0], h[2024][1], h[2025][0], h[2025][1],
        row["median"], row["best"], row["worst"], row["volatility"],
        row["reachable_years"], row["years"], judgement,
        "适合作为冲刺或波动机会，不可替代稳保志愿；核验招生计划、专业改名及校区变化",
    ])

for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="C65911")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    row[18].fill = PatternFill("solid", fgColor="FCE4D6")
widths = [8, 15, 27, 40, 12, 14, 12, 14, 12, 14, 12, 14, 16, 13, 13, 13, 15, 12, 25, 55]
for i, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width
ws.freeze_panes = "E2"
if results:
    table = Table(displayName="VolatileGoodSchools", ref=f"A1:T{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)

note = wb["学科标注说明"]
note.append(("波动机会表", "筛选双一流/原211及部分行业强校；至少有3年同校同名专业数据，位次波动不少于6000名，且本人位次位于波动区间或距区间不超过5000名。"))

wb.save(BOOK)
print(BOOK)
print("volatile rows:", len(results), "schools:", len({r["school_key"] for r in results}))
