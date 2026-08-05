import re
import statistics
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).parent
DATA = ROOT / "data"
TARGET_RANK = 26954

SOURCES = {
    2022: "https://cdn.gaokzx.com/zixunzhan/202403/a8849ca2-9596-4d78-8eed-a9c3ac89df49.pdf",
    2023: "https://cdn.gaokzx.com/zixunzhan/202403/88fefb53-5c97-4a16-bfa2-565abcbe37dd.pdf",
    2024: "https://cdn.zizzs.com/1721564943406%E7%89%A9%E7%90%86.pdf",
    2025: "https://cdn.gaokzx.com/zixunzhan/1753323564725%E7%89%A9%E7%90%86.pdf",
}
RANK_SOURCES = {
    2022: "https://www.dxsbb.com/news/117895.html",
    2023: "https://www.dxsbb.com/news/136970.html",
    2024: "https://www.dxsbb.com/news/146467.html",
    2025: "https://www.dxsbb.com/news/148772.html",
}

PROVINCES = {
    "11": "北京", "12": "天津", "13": "河北", "14": "山西", "15": "内蒙古",
    "21": "辽宁", "22": "吉林", "23": "黑龙江", "31": "上海", "32": "江苏",
    "33": "浙江", "34": "安徽", "35": "福建", "36": "江西", "37": "山东",
    "41": "河南", "42": "湖北", "43": "湖南", "44": "广东", "45": "广西",
    "46": "海南", "50": "重庆", "51": "四川", "52": "贵州", "53": "云南",
    "54": "西藏", "61": "陕西", "62": "甘肃", "63": "青海", "64": "宁夏",
    "65": "新疆",
}

GOOD = {
    "计算机": 11, "软件工程": 10, "电气工程": 10, "电子信息": 9, "自动化": 9,
    "通信工程": 9, "人工智能": 9, "数据科学": 8, "信息安全": 8, "物联网": 7,
    "机械": 6, "能源": 6, "新能源": 7, "数学": 6, "统计": 6, "交通": 6,
    "临床医学": 8, "口腔医学": 9, "医学影像": 7, "药学": 5, "法学": 5,
    "会计": 4, "财务": 4, "地理": 5, "测绘": 5, "遥感": 7, "材料": 3,
    "化学": 4, "生物医学工程": 7, "食品科学": 3, "水利": 5, "航空": 7,
}
BAD_WORDS = (
    "中外合作", "民族班", "预科班", "地方专项", "国家专项", "护理学",
    "学前教育", "旅游管理", "酒店管理", "土木工程", "园林", "农学",
    "职业技术大学", "独立学院",
)


def clean_school(name):
    name = re.sub(r"[（(](地方专项|国家专项|中外合作|民族班|预科班).*?[）)]", "", name)
    return name.strip()


def clean_major(name):
    name = re.sub(r"[（(].*?[）)]", "", name)
    name = re.sub(r"(类|试验班.*)$", "", name)
    return name.strip()


def parse_admissions(year):
    rows = []
    for raw in (DATA / f"{year}.txt").read_text(encoding="utf-8").splitlines():
        line = " ".join(raw.split())
        m = re.match(r"^(\d{4})\s+(.+?)\s+([0-9A-Z]{3})\s+(.+?)\s+(\d{3})(?:\s|$)", line)
        if not m:
            continue
        code, school, major_code, major, score = m.groups()
        score = int(score)
        if 350 <= score <= 750:
            rows.append({
                "year": year, "code": code, "school": school, "major_code": major_code,
                "major": major, "score": score, "school_key": clean_school(school),
                "major_key": clean_major(major),
            })
    return rows


def load_rank_map(year):
    table = pd.read_html(RANK_SOURCES[year])[0]
    table.columns = ["score", "count", "rank"]
    out = {}
    for _, row in table.iterrows():
        m = re.search(r"\d+", str(row["score"]))
        if m:
            try:
                out[int(m.group())] = int(float(row["rank"]))
            except (ValueError, TypeError):
                pass
    return out


def desirability(major):
    value = max((weight for word, weight in GOOD.items() if word in major), default=1)
    if any(word in major for word in ("智能", "信息", "工程")):
        value += 1
    return value


rank_maps = {year: load_rank_map(year) for year in range(2022, 2026)}
all_rows = {year: parse_admissions(year) for year in range(2022, 2026)}
indexes = {}
for year, rows in all_rows.items():
    idx = {}
    for row in rows:
        idx.setdefault((row["school_key"], row["major_key"]), []).append(row)
    indexes[year] = idx

candidates = []
for current in all_rows[2025]:
    if any(word in current["school"] or word in current["major"] for word in BAD_WORDS):
        continue
    if not (490 <= current["score"] <= 595):
        continue
    history = {}
    for year in range(2022, 2026):
        matches = indexes[year].get((current["school_key"], current["major_key"]), [])
        if matches:
            # 同校同专业出现多个招生类型时，取普通类型名称最短的一项。
            match = sorted(matches, key=lambda r: (len(r["school"]), len(r["major"])))[0]
            history[year] = (match["score"], rank_maps[year].get(match["score"]))
        else:
            history[year] = (None, None)
    ranks = [rank for _, rank in history.values() if rank]
    if len(ranks) < 3:
        continue
    med = round(statistics.median(ranks))
    if not (13500 <= med <= 50000):
        continue
    candidates.append({
        **current, "history": history, "median_rank": med,
        "score_desire": desirability(current["major"]),
        "years": len(ranks), "volatility": max(ranks) - min(ranks),
    })

# 每校优先保留两个专业，兼顾专业质量、数据完整度和与目标位次的接近程度。
candidates.sort(key=lambda x: (
    -x["score_desire"], -x["years"], abs(x["median_rank"] - TARGET_RANK), x["volatility"]
))
per_school = {}
diverse = []
for c in candidates:
    key = c["school_key"]
    if per_school.get(key, 0) >= 2:
        continue
    per_school[key] = per_school.get(key, 0) + 1
    diverse.append(c)

segments = {
    # 故意拉开位次跨度，避免100条都挤在考生位次附近，失去“冲稳保”的意义。
    "冲": [(13500, 18500, 8), (18500, 21000, 10), (21000, 23500, 12)],
    "稳": [(23500, 27000, 20), (27000, 30000, 15), (30000, 32501, 10)],
    "保": [(32501, 38000, 10), (38000, 44000, 8), (44000, 50001, 7)],
}
targets = {"冲": 30, "稳": 45, "保": 25}
selected = []
for band in ("冲", "稳", "保"):
    band_items = []
    for low, high, count in segments[band]:
        midpoint = (low + high) / 2
        pool = [c for c in diverse if low <= c["median_rank"] < high]
        pool.sort(key=lambda x: (
            -x["score_desire"], -x["years"], x["volatility"],
            abs(x["median_rank"] - midpoint),
        ))
        band_items.extend(pool[:count])
    band_items.sort(key=lambda x: x["median_rank"])
    selected.extend((band, item) for item in band_items)

if len(selected) != 100:
    raise RuntimeError(f"候选数量不足：仅选出 {len(selected)} 条")

wb = Workbook()
ws = wb.active
ws.title = "100个志愿建议"
headers = [
    "建议序号", "梯度", "院校代码(2025)", "院校", "省份", "专业代码(2025)", "专业",
    "选科适配初筛", "2022最低分", "2022最低位次", "2023最低分", "2023最低位次",
    "2024最低分", "2024最低位次", "2025最低分", "2025最低位次",
    "4年位次中位数", "与本人位次差", "近年波动范围", "有效年份数", "填报提示",
]
ws.append(headers)
for i, (band, c) in enumerate(selected, 1):
    h = c["history"]
    hint = {
        "冲": "历史中位位次高于本人；可少量冲刺，不能当保底",
        "稳": "历史中位位次接近本人；需结合2026计划与专业热度排序",
        "保": "历史中位位次低于本人；仍须保留更低位次的最终兜底",
    }[band]
    ws.append([
        i, band, c["code"], c["school"], PROVINCES.get(c["code"][:2], "待核验"),
        c["major_code"], c["major"], "物理+化学通常可报；以2026招生计划为准",
        h[2022][0], h[2022][1], h[2023][0], h[2023][1],
        h[2024][0], h[2024][1], h[2025][0], h[2025][1],
        c["median_rank"], c["median_rank"] - TARGET_RANK, c["volatility"], c["years"], hint,
    ])

tab = Table(displayName="Recommendations", ref=f"A1:U{ws.max_row}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tab)
ws.freeze_panes = "I2"
ws.auto_filter.ref = f"A1:U{ws.max_row}"
for row in ws.iter_rows(min_row=2):
    color = {"冲": "F4CCCC", "稳": "FFF2CC", "保": "D9EAD3"}[row[1].value]
    row[1].fill = PatternFill("solid", fgColor=color)
    for cell in row:
        cell.alignment = Alignment(vertical="center", wrap_text=True)
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
widths = [10, 7, 15, 24, 9, 16, 34, 25, 12, 14, 12, 14, 12, 14, 12, 14, 15, 13, 13, 12, 34]
for i, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[1].height = 34

guide = wb.create_sheet("使用说明")
notes = [
    ("项目", "说明"),
    ("考生", "2026重庆高考；物理+化学+地理；566分；位次26954"),
    ("数据口径", "2022—2025重庆普通类本科批、物理类、平行志愿专业投档最低分。最低投档分不等于录取概率。"),
    ("筛选方法", "以四年可匹配数据的位次中位数为主，按院校+专业形成100条建议；每校最多保留2个专业以增加院校覆盖。"),
    ("冲稳保定义", "冲：中位位次13500—23499；稳：23500—32500；保：32501—50000。是辅助分层，不是录取承诺。"),
    ("缺失值", "空白代表该年度未检出同名院校+同名专业，常见原因包括停招、新增、专业改名、招生类型变化。未用估算值填补。"),
    ("选科提醒", "物化地覆盖大多数理工农医专业，但2026实际选科、体检、单科、外语语种、性别等要求必须逐条核对招生计划。"),
    ("年份变化", "2026招生计划尚可能变化；专业代码每年可能重编，本表代码仅用于定位2025原始数据，不能直接用于2026填报。"),
    ("排序建议", "正式填报时应按真实喜好排序，并增加本人可接受的更低位次兜底项；不接受的专业不要仅因“保”而填。"),
    ("特别提醒", "表中未主动推荐中外合作、民族班、预科班、专项计划、护理学及职业本科；如你符合专项资格或接受高学费，可另做一版。"),
]
for row in notes:
    guide.append(row)
guide.column_dimensions["A"].width = 18
guide.column_dimensions["B"].width = 110
guide.freeze_panes = "A2"
for cell in guide[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
for row in guide.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

src = wb.create_sheet("数据来源")
src.append(["年份", "资料类型", "来源链接", "备注"])
for year in range(2022, 2026):
    src.append([year, "专业投档表", SOURCES[year], "重庆市教育考试院发布数据的PDF镜像"])
    src.append([year, "物理类一分一段表", RANK_SOURCES[year], "用于把最低分转换为当年累计位次"])
src.append(["2026", "考生输入", "", "566分、26954名、物化地"])
for cell in src[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
for row in range(2, src.max_row + 1):
    if src.cell(row, 3).value:
        src.cell(row, 3).hyperlink = src.cell(row, 3).value
        src.cell(row, 3).style = "Hyperlink"
src.column_dimensions["A"].width = 10
src.column_dimensions["B"].width = 22
src.column_dimensions["C"].width = 90
src.column_dimensions["D"].width = 45
src.freeze_panes = "A2"

detail = wb.create_sheet("入选数据长表")
detail.append(["建议序号", "梯度", "院校", "专业", "年份", "最低分", "最低位次"])
for i, (band, c) in enumerate(selected, 1):
    for year in range(2022, 2026):
        detail.append([i, band, c["school"], c["major"], year, *c["history"][year]])
detail.freeze_panes = "A2"
for cell in detail[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
for col, width in zip("ABCDEFG", [10, 8, 26, 36, 10, 12, 14]):
    detail.column_dimensions[col].width = width

out = ROOT / "2026重庆高考_物化地_566分26954名_100个志愿建议.xlsx"
wb.save(out)
print(out)
print({band: sum(1 for b, _ in selected if b == band) for band in targets})
print("distinct schools:", len({c["school_key"] for _, c in selected}))
