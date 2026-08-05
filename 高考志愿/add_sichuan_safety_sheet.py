import re
import statistics
import time
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).parent
BOOK = ROOT / "2026重庆高考_提前批AB段加本科批_整体填报方案.xlsx"
TARGET = 26954
RANK_URLS = {
    2022: "https://www.dxsbb.com/news/117895.html",
    2023: "https://www.dxsbb.com/news/136970.html",
    2024: "https://www.dxsbb.com/news/146467.html",
    2025: "https://www.dxsbb.com/news/148772.html",
}
PRIVATE = {
    "成都东软学院", "成都艺术职业大学", "电子科技大学成都学院", "成都理工大学工程技术学院",
    "四川传媒学院", "成都银杏酒店管理学院", "成都文理学院", "四川工商学院",
    "四川外国语大学成都学院", "四川工业科技学院", "成都锦城学院",
    "西南财经大学天府学院", "四川大学锦江学院", "四川文化艺术学院",
    "绵阳城市学院", "西南交通大学希望学院", "四川电影电视学院", "吉利学院",
}
FEATURES = {
    "西南交通大学": "交通运输、土木、电气、轨道交通",
    "西南财经大学": "应用经济、金融、工商管理",
    "四川农业大学": "农业科学、动物科学、食品科学",
    "成都理工大学": "地质资源、地球物理、土木、核技术",
    "西南石油大学": "石油天然气、化工、机械",
    "成都信息工程大学": "大气科学、信息通信、计算机",
    "西华大学": "机械、车辆、动力工程、食品科学",
    "四川师范大学": "教育学、中国语言文学、数学",
    "西南科技大学": "材料、环境、控制、土木",
    "四川轻化工大学": "化学工程、食品、材料、酿酒",
    "成都中医药大学": "中医学、中药学、针灸推拿",
    "西南医科大学": "临床医学、药学、基础医学",
    "川北医学院": "临床医学、医学影像、口腔医学",
    "成都大学": "机械、食品、计算机、临床医学",
    "中国民用航空飞行学院": "交通运输、航空运行、飞行技术",
}


def rank_map(year):
    last = None
    for attempt in range(5):
        try:
            table = pd.read_html(RANK_URLS[year])[0]
            break
        except Exception as exc:
            last = exc
            time.sleep(attempt + 1)
    else:
        raise last
    table.columns = ["score", "count", "rank"]
    out = {}
    for _, row in table.iterrows():
        m = re.search(r"\d+", str(row["score"]))
        if m:
            try:
                out[int(m.group())] = int(float(row["rank"]))
            except (TypeError, ValueError):
                pass
    return out


def clean_school(name):
    return re.sub(r"\((地方专项|国家专项|高校专项|中外合作|民族班|预科班).*?\)", "", name).strip()


def clean_major(name):
    return re.sub(r"\(.*?\)", "", name).removesuffix("类").strip()


ranks = {year: rank_map(year) for year in range(2022, 2026)}
year_rows = {}
for year in range(2022, 2026):
    data = []
    for raw in (ROOT / "data" / f"{year}.txt").read_text(encoding="utf-8").splitlines():
        line = " ".join(raw.split())
        m = re.match(r"^(\d{4})\s+(.+?)\s+([0-9A-Z]{3})\s+(.+?)\s+(\d{3})(?:\s|$)", line)
        if not m:
            continue
        code, school, major_code, major, score = m.groups()
        # 重庆招生院校代码前两位51对应四川院校；仅保留普通类型。
        if not code.startswith("51") or any(x in school for x in (
            "地方专项", "国家专项", "高校专项", "民族班", "预科班", "中外合作"
        )):
            continue
        score = int(score)
        rank = ranks[year].get(score)
        if rank:
            data.append({
                "code": code, "school": clean_school(school), "major_code": major_code,
                "major": major, "major_key": clean_major(major), "score": score, "rank": rank,
            })
    year_rows[year] = data

index = {}
for year, rows in year_rows.items():
    for row in rows:
        index.setdefault((year, row["school"], row["major_key"]), []).append(row)

results, seen = [], set()
for current in year_rows[2025]:
    key = (current["school"], current["major_key"])
    if key in seen or "思想政治教育" in current["major"]:
        continue
    seen.add(key)
    history = {}
    for year in range(2022, 2026):
        matches = index.get((year, *key), [])
        if matches:
            match = sorted(matches, key=lambda r: len(r["major"]))[0]
            history[year] = (match["score"], match["rank"])
        else:
            history[year] = (None, None)
    valid = [rank for _, rank in history.values() if rank]
    if len(valid) < 2:
        continue
    median = round(statistics.median(valid))
    private = current["school"] in PRIVATE
    if not private and median < 30000 and current["rank"] < 33000:
        continue
    if private:
        category = "民办最终兜底"
    elif median < 38000:
        category = "准保底（仍有波动）"
    elif median < 55000:
        category = "公办稳妥保底"
    else:
        category = "公办深度保底"
    results.append({
        **current, "history": history, "median": median, "best": min(valid),
        "worst": max(valid), "volatility": max(valid) - min(valid), "years": len(valid),
        "category": category, "private": private,
    })

order = {"准保底（仍有波动）": 1, "公办稳妥保底": 2, "公办深度保底": 3, "民办最终兜底": 4}
results.sort(key=lambda r: (order[r["category"]], r["median"], r["school"], r["major"]))

wb = load_workbook(BOOK)
if "四川院校保底" in wb.sheetnames:
    del wb["四川院校保底"]
ws = wb.create_sheet("四川院校保底", 4)
headers = [
    "序号", "保底层级", "办学性质", "院校代码", "院校", "专业", "院校特色方向",
    "2022最低分", "2022最低位次", "2023最低分", "2023最低位次",
    "2024最低分", "2024最低位次", "2025最低分", "2025最低位次",
    "近年位次中位数", "最好位次", "最差位次", "波动幅度", "有效年份",
    "相对本人位次差", "填报提示",
]
ws.append(headers)
for i, row in enumerate(results, 1):
    h = row["history"]
    nature = "民办" if row["private"] else "公办（以2026计划核验）"
    feature = FEATURES.get(row["school"], "应用型或区域特色方向，具体优势以学校官网为准")
    tip = "核验2026选科、校区、体检及专业培养方向"
    if row["private"]:
        tip = "最终兜底；重点核验学费、住宿、校区、转专业和培养模式"
    ws.append([
        i, row["category"], nature, row["code"], row["school"], row["major"], feature,
        h[2022][0], h[2022][1], h[2023][0], h[2023][1],
        h[2024][0], h[2024][1], h[2025][0], h[2025][1],
        row["median"], row["best"], row["worst"], row["volatility"], row["years"],
        row["median"] - TARGET, tip,
    ])

for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
colors = {
    "准保底（仍有波动）": "FFF2CC", "公办稳妥保底": "D9EAD3",
    "公办深度保底": "B6D7A8", "民办最终兜底": "D9D2E9",
}
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    row[1].fill = PatternFill("solid", fgColor=colors[row[1].value])
widths = [8, 20, 20, 12, 27, 40, 36, 12, 14, 12, 14, 12, 14, 12, 14, 16, 13, 13, 13, 12, 17, 50]
for i, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width
ws.freeze_panes = "H2"
if results:
    table = Table(displayName="SichuanSafetyChoices", ref=f"A1:V{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)

note = wb["学科标注说明"]
note.append(("四川保底表", "按重庆物理类近年专业投档位次分层。四川院校数量多，表内不限制专业数量；准保底不等于稳录，正式填报仍须保留更低位次项目。"))
wb.save(BOOK)
print(BOOK)
print("rows:", len(results), Counter(r["category"] for r in results))
