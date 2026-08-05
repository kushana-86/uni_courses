import os
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


target = Path(os.environ["TARGET_XLSX"])
workbook = load_workbook(target)
sheet = workbook.worksheets[4]

# The source sheet has no header: A/C/D/F contain the original volunteer
# number, university, ownership, and major respectively.
records = []
for row in sheet.iter_rows(values_only=True):
    if not any(value is not None for value in row):
        continue
    records.append(
        {
            "original_order": row[0],
            "university": str(row[2] or "").strip(),
            "ownership": row[3],
            "major": str(row[5] or "").strip(),
        }
    )

school_header = "\u9662\u6821"
major_header = "\u4e13\u4e1a"
rank_header = "2025\u6700\u4f4e\u4f4d\u6b21"
rank_index = {}

for source in target.parent.parent.rglob("*.xlsx"):
    if source.resolve() == target.resolve():
        continue
    try:
        source_book = load_workbook(source, read_only=True, data_only=True)
    except Exception:
        continue
    for source_sheet in source_book.worksheets:
        rows = source_sheet.iter_rows(values_only=True)
        try:
            headers = [str(value or "").strip() for value in next(rows)]
        except StopIteration:
            continue
        if not all(name in headers for name in (school_header, major_header, rank_header)):
            continue
        school_col = headers.index(school_header)
        major_col = headers.index(major_header)
        rank_col = headers.index(rank_header)
        for row in rows:
            if len(row) <= max(school_col, major_col, rank_col):
                continue
            rank = row[rank_col]
            if not isinstance(rank, (int, float)):
                continue
            key = (
                str(row[school_col] or "").strip(),
                str(row[major_col] or "").strip(),
            )
            rank_index.setdefault(key, int(rank))

for record in records:
    record["rank"] = rank_index.get((record["university"], record["major"]))

records.sort(
    key=lambda record: (
        record["rank"] is None,
        record["rank"] if record["rank"] is not None else float("inf"),
        record["university"],
        record["major"],
    )
)

sheet.delete_rows(1, sheet.max_row)
headers = [
    "\u6392\u5e8f",
    "\u539f\u5fd7\u613f\u5e8f\u53f7",
    "\u9662\u6821",
    "\u6027\u8d28",
    "\u4e13\u4e1a",
    "2025\u6700\u4f4e\u4f4d\u6b21",
    "\u6570\u636e\u72b6\u6001",
]
sheet.append(headers)

for index, record in enumerate(records, start=1):
    has_rank = record["rank"] is not None
    sheet.append(
        [
            index,
            record["original_order"],
            record["university"],
            record["ownership"],
            record["major"],
            record["rank"] if has_rank else "\u65e02025\u6570\u636e",
            "\u5df2\u5339\u914d" if has_rank else "\u65e02025\u6570\u636e\uff0c\u5df2\u7f6e\u4e8e\u672b\u5c3e",
        ]
    )

header_fill = PatternFill("solid", fgColor="1F4E78")
missing_fill = PatternFill("solid", fgColor="FFF2CC")
for cell in sheet[1]:
    cell.fill = header_fill
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row in sheet.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="center")
    if row[5].value == "\u65e02025\u6570\u636e":
        for cell in row:
            cell.fill = missing_fill

# Reset any legacy desktop-view pane before freezing only the header row.
# Merely assigning A2 can leave an old pane/selection state in some Excel files.
sheet.freeze_panes = None
sheet.sheet_view.selection = []
sheet.sheet_view.topLeftCell = "A1"
sheet.freeze_panes = "A2"
sheet.auto_filter.ref = f"A1:G{sheet.max_row}"
sheet.column_dimensions["A"].width = 8
sheet.column_dimensions["B"].width = 14
sheet.column_dimensions["C"].width = 24
sheet.column_dimensions["D"].width = 10
sheet.column_dimensions["E"].width = 38
sheet.column_dimensions["F"].width = 18
sheet.column_dimensions["G"].width = 28

workbook.save(target)
