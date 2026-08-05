import re
import time
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).parent
TARGET_RANK = 26954
RANK_URL = "https://www.dxsbb.com/news/148772.html"

A_SCHOOLS = [
    ("A段", "军校", "陆军防化学院", 562, "男生最低线；须军检、政审、面试，专业组另核"),
    ("A段", "军校", "武警警官学院", 565, "男生最低线；须军检、政审、面试"),
    ("A段", "军校", "海军大连舰艇学院", 567, "男生最低线；须军检、政审、面试"),
    ("A段", "军校", "陆军军医大学", 577, "须军检、政审、面试；体检标准严格"),
    ("A段", "军校", "陆军兵种大学", 581, "男生最低线；须军检、政审、面试"),
    ("A段", "军校", "陆军步兵学院", 582, "男生最低线；须军检、政审、面试"),
    ("A段", "军校", "火箭军工程大学", 585, "男生最低线；须军检、政审、面试"),
    ("A段", "军校", "军事航天部队航天工程大学", 585, "男生最低线；须军检、政审、面试"),
    ("A段", "军校", "武警工程大学", 587, "男生最低线；须军检、政审、面试"),
    ("A段", "军校", "空军预警学院", 588, "男生最低线；须军检、政审、面试"),
    ("A段", "军校", "陆军工程大学", 594, "男生最低线；须军检、政审、面试"),
    ("A段", "军校", "空军工程大学", 595, "男生最低线；须军检、政审、面试"),
    ("A段", "军校", "联勤保障部队工程大学", 602, "男生最低线；须军检、政审、面试"),
    ("A段", "军校", "海军工程大学", 603, "男生最低线；须军检、政审、面试"),
    ("A段", "军校", "海军航空大学", 602, "男生最低线；须军检、政审、面试"),
    ("A段", "军校", "信息支援部队工程大学", 621, "男生最低线；须军检、政审、面试"),
    ("A段", "军校", "网络空间部队信息工程大学", 627, "男生最低线；须军检、政审、面试"),
    ("A段", "军校", "国防科技大学", 638, "物理类男生最低线；女生最低线更高"),
    ("A段", "警校", "南京警察学院", 480, "学校最低线含不同性别/专业；须公安政审、面试、体检、体测"),
    ("A段", "警校", "重庆警察学院", 564, "学校最低线；须公安政审、面试、体检、体测"),
    ("A段", "警校", "中央司法警官学院", 571, "须政审、面试、体检、体测；专业选科另核"),
    ("A段", "警校", "中国人民警察大学", 571, "须公安政审、面试、体检、体测"),
    ("A段", "警校", "郑州警察学院", 572, "须公安政审、面试、体检、体测"),
    ("A段", "警校", "中国刑事警察学院", 609, "须公安政审、面试、体检、体测"),
    ("A段", "警校", "中国人民公安大学", 620, "物理类男生学校最低线；女生更高"),
]


def load_rank_map():
    last = None
    for attempt in range(5):
        try:
            table = pd.read_html(RANK_URL)[0]
            break
        except Exception as exc:
            last = exc
            time.sleep(attempt + 1)
    else:
        raise last
    table.columns = ["score", "count", "rank"]
    result = {}
    for _, row in table.iterrows():
        m = re.search(r"\d+", str(row["score"]))
        if m:
            try:
                result[int(m.group())] = int(float(row["rank"]))
            except (ValueError, TypeError):
                pass
    return result


rank_map = load_rank_map()


def parse_b():
    rows = []
    for raw in (ROOT / "data" / "2025_B_layout.txt").read_text(encoding="utf-8").splitlines():
        line = " ".join(raw.split())
        m = re.match(r"^(\d{4})\s*(.+?)\s+([0-9A-Z]{3})\s*(.+?)\s*(\d{3})(?:\s|$)", line)
        if not m:
            continue
        code, school, major_code, major, score = m.groups()
        score = int(score)
        rank = rank_map.get(score)
        if not rank or not (14000 <= rank <= 52000):
            continue
        text = school + major
        if "国家专项" in text:
            kind, qualification = "国家专项", "仅限通过国家专项资格审核"
        elif "地方专项" in text:
            kind, qualification = "地方专项", "仅限通过地方专项资格审核"
        elif "免费医学定向" in text:
            kind, qualification = "免费医学定向", "须符合定向区县条件并签约"
        elif any(x in major for x in ("航海技术", "轮机工程", "船舶电子电气工程")):
            kind, qualification = "航海类", "须核验视力、色觉、身体条件和性别要求"
        else:
            kind, qualification = "其他B段", "核验2026计划中的资格与身体条件"
        rows.append({
            "code": code, "school": school, "major_code": major_code, "major": major,
            "score": score, "rank": rank, "kind": kind, "qualification": qualification,
        })
    return rows


def major_value(text):
    weights = {
        "计算机": 12, "电气": 11, "电子信息": 10, "人工智能": 10, "软件": 10,
        "数据科学": 9, "自动化": 9, "临床医学": 9, "医学影像": 8, "数学": 7,
        "统计": 7, "智能制造": 7, "交通运输": 7, "水利": 6, "机械": 6,
        "新能源": 7, "测绘": 5, "地理": 5, "法学": 5, "会计": 4,
    }
    return max((v for k, v in weights.items() if k in text), default=2)


b_all = parse_b()
b_selected, per_school = [], {}
band_plan = [(14000, 20000, 5), (20000, 23954, 10), (23954, 29955, 20),
             (29955, 40000, 10), (40000, 52001, 5)]
for low, high, count in band_plan:
    pool = [r for r in b_all if low <= r["rank"] < high]
    pool.sort(key=lambda r: (-major_value(r["major"]), abs(r["rank"] - TARGET_RANK)))
    added = 0
    for row in pool:
        school_key = re.sub(r"\(.*?\)", "", row["school"])
        if per_school.get(school_key, 0) >= 3:
            continue
        per_school[school_key] = per_school.get(school_key, 0) + 1
        b_selected.append(row)
        added += 1
        if added == count:
            break
b_selected.sort(key=lambda r: r["rank"])

wb = Workbook()
overview = wb.active
overview.title = "总体填报方案"
overview_rows = [
    ("录取顺序", "板块", "建议数量", "本人成绩下的定位", "执行方案", "未满足条件时"),
    (1, "本科提前批A段", "条件式填报", "军校/警校中少量项目接近566分", "仅在本人自愿且已通过政审、面试、体检/体测时填；优先按真实职业意愿排序", "整段放弃，不影响后续B段和本科批"),
    (2, "本科提前批B段", "建议表50项中择取", "国家专项、地方专项、免费医学定向、航海类需分别核验", "有专项资格才填专项；接受定向服务才填医学定向；身体条件合格才填航海", "删除不具资格项目，保留合格项目"),
    (3, "普通本科批", "100个院校+专业", "主战场：冲30、稳45、保25", "按专业接受度重排；不接受的专业不填；最后保留真正兜底", "这是最终主要保障，不能因填提前批而草率"),
    (4, "非西藏生定藏就业", "18项条件式备选", "重庆文理学院3个专业与本人位次最接近", "只有明确接受赴藏定向就业、服务年限和违约条款时才填", "不接受定向就业则全部删除"),
]
for row in overview_rows:
    overview.append(row)


def style_sheet(ws, widths, color="1F4E78", freeze="A2", table_name=None):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = freeze
    if table_name and ws.max_row > 1:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        tab = Table(displayName=table_name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tab)


style_sheet(overview, [12, 22, 20, 38, 65, 42], color="375623")

a_ws = wb.create_sheet("A段条件式备选")
a_ws.append(["序号", "类型", "院校", "2025学校最低分", "2025对应位次", "与本人位次差", "初步定位", "资格与风险"])
for i, (_, kind, school, score, condition) in enumerate(A_SCHOOLS, 1):
    rank = rank_map.get(score)
    diff = rank - TARGET_RANK if rank else None
    if rank is None:
        label = "待核验"
    elif diff < -3000:
        label = "偏冲"
    elif diff <= 3000:
        label = "接近"
    else:
        label = "分数层面可够；资格决定"
    a_ws.append([i, kind, school, score, rank, diff, label, condition])
style_sheet(a_ws, [8, 10, 28, 16, 16, 16, 22, 60], color="C00000", table_name="AChoices")

b_ws = wb.create_sheet("B段建议50项")
b_ws.append(["序号", "类型", "院校代码", "院校", "专业代码", "专业", "2025最低分", "2025最低位次", "与本人位次差", "初步定位", "资格条件"])
for i, row in enumerate(b_selected, 1):
    diff = row["rank"] - TARGET_RANK
    label = "偏冲" if diff < -3000 else ("接近" if diff <= 3000 else "偏稳")
    b_ws.append([i, row["kind"], row["code"], row["school"], row["major_code"], row["major"],
                 row["score"], row["rank"], diff, label, row["qualification"]])
style_sheet(b_ws, [8, 16, 12, 30, 12, 42, 14, 15, 16, 12, 40], color="BF9000", table_name="BChoices")

# 将已经制作的本科批100项原样并入总方案。
main_file = ROOT / "2026重庆高考_物化地_566分26954名_100个志愿建议.xlsx"
main_book = load_workbook(main_file, data_only=True)
main_source = main_book["100个志愿建议"]
main_ws = wb.create_sheet("普通本科批100项")
for row in main_source.iter_rows(values_only=True):
    main_ws.append(list(row))
style_sheet(main_ws, [10, 7, 15, 24, 9, 16, 34, 25, 12, 14, 12, 14, 12, 14, 12, 14, 15, 13, 13, 12, 34],
            color="1F4E78", freeze="I2", table_name="MainChoices")

tibet_file = ROOT / "重庆物理类_非西藏生定藏就业_筛选表.xlsx"
tibet_book = load_workbook(tibet_file, data_only=True)
tibet_source = tibet_book["2025定向西藏计划"]
tibet_ws = wb.create_sheet("定藏就业条件备选")
for row in tibet_source.iter_rows(values_only=True):
    tibet_ws.append(list(row))
style_sheet(tibet_ws, [8, 12, 24, 12, 36, 25, 13, 15, 18, 12, 13, 15, 13, 15, 13, 15, 55],
            color="7030A0", freeze="F2", table_name="TibetChoices")

check = wb.create_sheet("填报前资格核对")
checks = [
    ("核对项", "是/否（自行填写）", "影响板块", "说明"),
    ("是否愿意报军校并接受军队职业安排", "", "A段军校", "不愿意则删除全部军校"),
    ("是否完成并通过军校政审、军检、面试", "", "A段军校", "任一未通过均不能填"),
    ("是否愿意报公安/司法警校", "", "A段警校", "须接受公安司法类培养与就业方向"),
    ("是否完成相应政审、面试、体检、体测", "", "A段警校", "不同院校程序不同"),
    ("是否通过国家专项资格审核", "", "B段国家专项", "未通过则删除全部国家专项"),
    ("是否通过地方专项资格审核", "", "B段地方专项", "未通过则删除全部地方专项"),
    ("是否接受免费医学定向服务和签约", "", "B段免费医学定向", "同时核验定向区县资格"),
    ("是否符合航海类视力、色觉等要求", "", "B段航海类", "不符合则删除航海类"),
    ("是否接受赴西藏定向就业及违约条款", "", "定藏就业", "不接受则删除全部定藏就业项目"),
    ("是否逐项核对2026选科、体检、语种要求", "", "全部", "2025数据不能替代2026招生计划"),
]
for row in checks:
    check.append(row)
style_sheet(check, [38, 22, 24, 70], color="44546A")

sources = wb.create_sheet("数据来源与说明")
source_rows = [
    ("资料", "链接/说明"),
    ("2025本科提前批B段物理类投档表", "https://cdn.gaokzx.com/zixunzhan/1753263227044%E9%87%8D%E5%BA%86%E5%B8%82%E6%99%AE%E9%80%9A%E9%AB%98%E6%A0%A1%E6%8B%9B%E7%94%9F%E4%BF%A1%E6%81%AF%E8%A1%A8%E6%9C%AC%E7%A7%91%E6%8F%90%E5%89%8D%E6%89%B9B%E6%AE%B5-%E7%89%A9%E7%90%86%E7%B1%BB.pdf"),
    ("2025物理类一分一段表", RANK_URL),
    ("A段军警最低分汇总", "https://www.gk100.com/read_9894173.htm"),
    ("重要口径", "A段表采用学校/性别类别最低分作初筛，不等于具体专业可录取；B段采用专业投档最低分。"),
    ("2026提醒", "2026正式招生计划、专业代码、资格要求尚须以重庆市教育考试院和高校招生章程为准。"),
]
for row in source_rows:
    sources.append(row)
for r in range(2, sources.max_row + 1):
    value = sources.cell(r, 2).value
    if isinstance(value, str) and value.startswith("http"):
        sources.cell(r, 2).hyperlink = value
        sources.cell(r, 2).style = "Hyperlink"
style_sheet(sources, [30, 115], color="44546A")

out = ROOT / "2026重庆高考_提前批AB段加本科批_整体填报方案.xlsx"
wb.save(out)
print(out)
print("A:", len(A_SCHOOLS), "B:", len(b_selected), "main:", main_ws.max_row - 1, "tibet:", tibet_ws.max_row - 1)
