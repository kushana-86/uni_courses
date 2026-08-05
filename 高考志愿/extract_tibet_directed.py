import re
import time
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from pypdf import PdfReader

ROOT = Path(__file__).parent
TARGET = 26954
NEEDLE = "定藏就业"
RANK_URLS = {
    2022: "https://www.dxsbb.com/news/117895.html",
    2023: "https://www.dxsbb.com/news/136970.html",
    2024: "https://www.dxsbb.com/news/146467.html",
    2025: "https://www.dxsbb.com/news/148772.html",
}


def rank_map(year):
    last_error = None
    for attempt in range(5):
        try:
            table = pd.read_html(RANK_URLS[year])[0]
            break
        except Exception as exc:
            last_error = exc
            time.sleep(1 + attempt)
    else:
        raise last_error
    table.columns = ["score", "count", "rank"]
    result = {}
    for _, row in table.iterrows():
        m = re.search(r"\d+", str(row["score"]))
        if m:
            try:
                result[int(m.group())] = int(float(row["rank"]))
            except (ValueError, TypeError):
                pass
    return result


def clean(text):
    text = text.replace("(非西藏生定藏就业)", "")
    text = re.sub(r"\(师范类\)", "", text)
    return text.strip()


def extract(year):
    reader = PdfReader(ROOT / "data" / f"{year}.pdf")
    rows = []
    for page in reader.pages:
        plain = page.extract_text() or ""
        if NEEDLE not in plain:
            continue
        text = page.extract_text(extraction_mode="layout") or plain
        for raw in text.splitlines():
            if NEEDLE not in raw:
                continue
            line = " ".join(raw.split())
            m = re.match(
                r"^(\d{4})\s*(.+?\(非西藏生定藏就业\))\s*([0-9A-Z]{3})\s*(.+?)\s+(\d{3})(?:\s|$)",
                line,
            )
            if not m:
                continue
            code, school, major_code, major, score = m.groups()
            rows.append({
                "year": year, "code": code, "school": school, "major_code": major_code,
                "major": major, "score": int(score), "school_key": clean(school),
                "major_key": clean(major),
            })
    return rows


ranks = {year: rank_map(year) for year in range(2022, 2026)}
records = []
for year in range(2022, 2026):
    for row in extract(year):
        row["rank"] = ranks[year].get(row["score"])
        records.append(row)

index = {}
for row in records:
    index.setdefault((row["year"], row["school_key"], row["major_key"]), row)

current = [row for row in records if row["year"] == 2025]
current.sort(key=lambda r: (r["rank"] or 999999, r["school_key"], r["major_key"]))

wb = Workbook()
ws = wb.active
ws.title = "2025定向西藏计划"
headers = [
    "序号", "院校代码", "院校", "专业代码", "专业", "物化地适配初筛", "2025最低分", "2025最低位次",
    "与本人26954名差值", "位置判断", "2022最低分", "2022最低位次",
    "2023最低分", "2023最低位次", "2024最低分", "2024最低位次", "重要提醒",
]
ws.append(headers)
for i, row in enumerate(current, 1):
    history = {}
    for year in (2022, 2023, 2024):
        history[year] = index.get((year, row["school_key"], row["major_key"]))
    diff = (row["rank"] - TARGET) if row["rank"] else None
    if diff is None:
        label = "位次缺失"
    elif diff < -3000:
        label = "偏冲"
    elif diff <= 3000:
        label = "位次接近"
    else:
        label = "偏稳"
    values = []
    for year in (2022, 2023, 2024):
        old = history[year]
        values.extend([old["score"] if old else None, old["rank"] if old else None])
    ws.append([
        i, row["code"], row["school_key"], row["major_code"], row["major_key"],
        "不适配：通常要求思想政治" if "思想政治" in row["major_key"] else "初筛适配；以2026计划为准",
        row["score"], row["rank"], diff, label, *values,
        "非西藏生源定向西藏就业；务必核验2026招生章程、签约、服务年限、就业地点及违约责任",
    ])

tab = Table(displayName="TibetDirected2025", ref=f"A1:Q{ws.max_row}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
ws.add_table(tab)
ws.freeze_panes = "F2"
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="7030A0")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    row[9].fill = PatternFill("solid", fgColor={
        "偏冲": "F4CCCC", "位次接近": "FFF2CC", "偏稳": "D9EAD3", "位次缺失": "D9D9D9"
    }[row[9].value])
widths = [8, 12, 24, 12, 36, 25, 13, 15, 18, 12, 13, 15, 13, 15, 13, 15, 55]
for i, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

raw = wb.create_sheet("2022-2025全部记录")
raw.append(["年份", "院校代码", "院校", "专业代码", "专业", "最低分", "最低位次"])
for row in sorted(records, key=lambda r: (r["year"], r["code"], r["major_code"])):
    raw.append([row["year"], row["code"], row["school_key"], row["major_code"],
                row["major_key"], row["score"], row["rank"]])
raw.freeze_panes = "A2"
for cell in raw[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="7030A0")
for col, width in zip("ABCDEFG", [10, 12, 26, 12, 38, 13, 15]):
    raw.column_dimensions[col].width = width

note = wb.create_sheet("说明")
notes = [
    ("项目", "说明"),
    ("筛选对象", "重庆普通类本科批物理类投档表中标注“非西藏生定藏就业”的院校专业。西藏地区高校普通招生不等于定向西藏就业，未混入本表。"),
    ("考生", "2026重庆物理类，物化地，566分，26954名。"),
    ("位置判断", "仅用2025最低位次辅助判断：比本人高3000名以上为偏冲，±3000名为接近，比本人低3000名以上为偏稳。"),
    ("历年空白", "代表同一院校同名专业在该年未检出，可能是停招、新增或改名，不以估算值替代。"),
    ("关键风险", "定向计划不是普通就业。填报前必须阅读2026招生章程和协议，确认培养、毕业去向、服务年限、户籍/资格、体检以及违约责任。"),
]
for item in notes:
    note.append(item)
note.column_dimensions["A"].width = 18
note.column_dimensions["B"].width = 115
for cell in note[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="7030A0")
for row in note.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

out = ROOT / "重庆物理类_非西藏生定藏就业_筛选表.xlsx"
wb.save(out)
print(out)
print("2025 rows:", len(current), "all records:", len(records))
